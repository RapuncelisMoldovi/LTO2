#!/usr/bin/env python3
"""
Импорт остатков на склад из Excel (без GUI).

Колонки на активном листе:
  A — номенклатурный номер (полный н/н варианта, как в справочнике)
  B — наименование (игнорируется для привязки; может быть неточным)
  C — единица измерения (не используется — берётся из номенклатуры)
  D — заводской номер
  E — год выпуска

Позиция на складе определяется только по н/н из вашей номенклатуры; отсутствующие
н/н в файле не создаются — сначала добавьте их в приложении.

Запуск из каталога репозитория (рядом с warehouse.db):
  python import_stock_from_excel.py путь\\к\\файлу.xls
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

from database import DatabaseManager

DEFAULT_DB_NAME = "warehouse.db"
DEFAULT_DOC = "Импорт остатков (Excel)"
DEFAULT_UNIT = "Склад-основной"


def _cell_str(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_year(value: object | None) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return int(value) if value.is_integer() else None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s.replace(",", ".")))
    except ValueError:
        return None


def _looks_like_header(col_a: str, col_b: str) -> bool:
    a = col_a.lower()
    b = col_b.lower()
    hints_a = (
        "номенклатур",
        "н/н",
        "нн",
        "код",
        "артикул",
    )
    hints_b = ("наименован", "название", "name", "изделие")
    return any(h in a for h in hints_a) and any(h in b for h in hints_b)


def _find_variant_by_full_code(db: DatabaseManager, full_code: str):
    cur = db.conn.cursor()
    cur.execute(
        """
        SELECT v.id, i.id AS item_id, i.name AS item_name, i.type AS item_type, i.uom
        FROM variants v
        JOIN items i ON i.id = v.item_id
        WHERE v.full_code = ?
        """,
        (full_code,),
    )
    return cur.fetchone()


def _resolve_serial_variant_by_full_code(
    db: DatabaseManager,
    full_code: str,
) -> tuple[int | None, str | None, str]:
    """
    Ищет вариант только по полному н/н (variants.full_code).

    Возвращает (variant_id, сообщение_об_ошибке_или_None, наименование_из_номенклатуры).
    """
    row = _find_variant_by_full_code(db, full_code)
    if row is None:
        return (
            None,
            "н/н отсутствует в номенклатуре — добавьте позицию в справочник и повторите импорт",
            "",
        )
    nomen_name = (row["item_name"] or "").strip()
    if row["item_type"] != "serial":
        return (
            None,
            f"ведётся количественный учёт, не серийный (в номенклатуре: «{nomen_name}»)",
            nomen_name,
        )
    return int(row["id"]), None, nomen_name


def import_stock_rows(
    db: DatabaseManager,
    rows: list[tuple[str, str, str, str, int | None]],
    *,
    unit_id: int | None,
    doc_name: str,
) -> tuple[int, int, list[str]]:
    """
    rows: (full_code, excel_name_ignored, uom_ignored, factory_sn, year)
    Возвращает (успешно, пропущено, список предупреждений/ошибок).
    """
    ok = 0
    skipped = 0
    messages: list[str] = []
    seen_variant_sn: set[tuple[int, str]] = set()
    name_mismatch_noted: set[str] = set()

    for idx, (base_code, excel_name, _uom, factory_sn, year) in enumerate(rows, start=1):
        line = f"Строка данных {idx}"
        if not factory_sn:
            messages.append(f"{line}: пустой заводской номер — пропуск")
            skipped += 1
            continue

        vid, err, nomen_name = _resolve_serial_variant_by_full_code(db, base_code)
        if err:
            messages.append(f"{line}: н/н «{base_code}» — {err}")
            skipped += 1
            continue

        sn_key = factory_sn.strip()
        vk = (vid, sn_key)
        if vk in seen_variant_sn:
            messages.append(
                f"{line}: повтор н/н «{base_code}» + S/N «{sn_key}» в файле — пропуск"
            )
            skipped += 1
            continue
        seen_variant_sn.add(vk)

        if db.serial_exists_for_variant(vid, sn_key):
            messages.append(
                f"{line}: S/N «{sn_key}» уже на складе по этому н/н — пропуск"
            )
            skipped += 1
            continue

        ex = excel_name.strip()
        if ex and nomen_name and ex.casefold() != nomen_name.casefold():
            if base_code not in name_mismatch_noted:
                name_mismatch_noted.add(base_code)
                messages.append(
                    f"{line}: наименование в файле («{ex}») не совпадает со справочником "
                    f"(«{nomen_name}») — использован н/н «{base_code}»"
                )

        basket = [
            {
                "variant_id": vid,
                "item_type": "serial",
                "sn": factory_sn,
                "manufacture_year": year,
                "qty": 1,
            }
        ]
        try:
            db.post_operation(basket, "IN", unit_id, doc_name, work_order_id=None)
            ok += 1
        except Exception as e:
            messages.append(f"{line}: не удалось провести приход «{factory_sn}»: {e}")
            skipped += 1

    return ok, skipped, messages


def _load_raw_rows_xlsx(path: Path) -> tuple[list[tuple[str, str, str, str, int | None]], str | None]:
    try:
        import openpyxl
    except ImportError:
        return [], "Установите openpyxl: pip install openpyxl"

    raw: list[tuple[str, str, str, str, int | None]] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return [], "В книге нет активного листа"

        for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
            if not row:
                continue
            col_a = _cell_str(row[0] if len(row) > 0 else None)
            col_b = _cell_str(row[1] if len(row) > 1 else None)
            if not col_a and not col_b:
                continue
            col_c = _cell_str(row[2] if len(row) > 2 else None)
            col_d = _cell_str(row[3] if len(row) > 3 else None)
            col_e = row[4] if len(row) > 4 else None
            year = _parse_year(col_e)
            raw.append((col_a, col_b, col_c, col_d, year))
        wb.close()
    except Exception as e:
        return [], f"Ошибка чтения файла: {e}"
    return raw, None


def _load_raw_rows_xls(path: Path) -> tuple[list[tuple[str, str, str, str, int | None]], str | None]:
    try:
        import xlrd
        from xlrd import XL_CELL_DATE
        from xlrd.xldate import xldate_as_datetime
    except ImportError:
        return [], "Установите xlrd: pip install xlrd"

    raw: list[tuple[str, str, str, str, int | None]] = []
    try:
        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            sheet = book.sheet_by_index(0)
        except xlrd.XLRDError:
            return [], "В книге нет листов"

        for r in range(sheet.nrows):
            vals: list[object | None] = []
            for c in range(5):
                if c >= sheet.ncols:
                    vals.append(None)
                    continue
                cell = sheet.cell(r, c)
                if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    vals.append(None)
                elif cell.ctype == XL_CELL_DATE:
                    try:
                        vals.append(xldate_as_datetime(cell.value, book.datemode))
                    except Exception:
                        vals.append(cell.value)
                else:
                    vals.append(cell.value)

            v0, v1, v2, v3, v4 = vals[0], vals[1], vals[2], vals[3], vals[4]
            col_a = _cell_str(v0)
            col_b = _cell_str(v1)
            if not col_a and not col_b:
                continue
            col_c = _cell_str(v2)
            col_d = _cell_str(v3)
            if isinstance(v4, datetime):
                year = v4.year
            else:
                year = _parse_year(v4)
            raw.append((col_a, col_b, col_c, col_d, year))
        book.release_resources()
    except Exception as e:
        return [], f"Ошибка чтения .xls: {e}"
    return raw, None


def read_excel(path: Path, *, skip_header: bool) -> tuple[list[tuple[str, str, str, str, int | None]], list[str]]:
    errors: list[str] = []
    out: list[tuple[str, str, str, str, int | None]] = []

    suf = path.suffix.lower()
    if suf == ".xls":
        raw, load_err = _load_raw_rows_xls(path)
    elif suf in (".xlsx", ".xlsm"):
        raw, load_err = _load_raw_rows_xlsx(path)
    else:
        return [], [
            f"Поддерживаются .xls, .xlsx и .xlsm; у файла расширение «{suf or '(нет)'}»"
        ]

    if load_err:
        return [], [load_err]

    if not raw:
        return [], ["В файле нет строк с данными в колонках A–E"]

    if skip_header and len(raw) >= 1:
        a0, b0 = raw[0][0], raw[0][1]
        if _looks_like_header(a0, b0):
            raw = raw[1:]

    for idx, (col_a, col_b, col_c, col_d, year) in enumerate(raw, start=1):
        base_code = col_a.strip()
        name = col_b.strip()
        uom = col_c.strip() or "шт"
        factory_sn = col_d.strip()
        if not base_code:
            errors.append(f"Строка {idx}: нет номенклатурного номера (колонка A)")
            continue
        out.append((base_code, name, uom, factory_sn, year))

    return out, errors


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Импорт серийных остатков из Excel в warehouse.db",
    )
    parser.add_argument(
        "excel",
        type=Path,
        help="Путь к .xls / .xlsx (A — н/н; D — зав. №; E — год)",
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=None,
        help=f"Путь к SQLite (по умолчанию: каталог скрипта / {DEFAULT_DB_NAME})",
    )
    parser.add_argument("--doc", type=str, default=DEFAULT_DOC, help="Номер/название документа прихода")
    parser.add_argument(
        "--unit",
        type=str,
        default=DEFAULT_UNIT,
        help="Подразделение для журнала (имя; по умолчанию склад основной)",
    )
    parser.add_argument(
        "--no-unit",
        action="store_true",
        help="Не подставлять подразделение (unit_id = NULL)",
    )
    parser.add_argument(
        "--no-skip-header",
        action="store_true",
        help="Не отбрасывать первую строку даже если она похожа на заголовок",
    )
    parser.add_argument("-q", "--quiet", action="store_true", help="Только итог и ошибки")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(message)s",
    )

    excel_path = args.excel.resolve()
    if not excel_path.is_file():
        print(f"Файл не найден: {excel_path}", file=sys.stderr)
        return 2

    script_dir = Path(__file__).resolve().parent
    db_path = args.db.resolve() if args.db else (script_dir / DEFAULT_DB_NAME)

    rows, read_errors = read_excel(excel_path, skip_header=not args.no_skip_header)
    for e in read_errors:
        print(e, file=sys.stderr)

    if not rows:
        if read_errors:
            return 1
        print("Нет строк для импорта.", file=sys.stderr)
        return 1

    with DatabaseManager(str(db_path)) as db:
        unit_id: int | None = None
        if not args.no_unit:
            unit_id = db.get_or_create_unit(args.unit.strip() or DEFAULT_UNIT)
        ok, skipped, messages = import_stock_rows(
            db,
            rows,
            unit_id=unit_id,
            doc_name=args.doc.strip() or DEFAULT_DOC,
        )

    for m in messages:
        print(m, file=sys.stderr)

    print(f"Готово: принято {ok}, пропущено {skipped}. БД: {db_path}")
    if not args.quiet and ok:
        print(f"Документ: {args.doc.strip() or DEFAULT_DOC}")
    return 0 if ok > 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
