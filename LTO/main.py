import sys
import os
import sqlite3
import json
import time
import logging
import shutil
from datetime import datetime

from PyQt6.QtCore import Qt, QSize, QDate, pyqtSignal
from PyQt6.QtGui import QIcon, QFontDatabase, QFont, QPixmap, QPainter, QColor
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QListWidget,
    QListWidgetItem,
    QPushButton,
    QLabel,
    QLineEdit,
    QTabWidget,
    QMessageBox,
    QDialog,
    QFormLayout,
    QComboBox,
    QRadioButton,
    QButtonGroup,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QTreeWidget,
    QTreeWidgetItem,
    QHeaderView,
    QFrame,
    QAbstractItemView,
    QInputDialog,
    QDateEdit,
    QFileDialog,
    QGridLayout,
    QToolButton,
)


DB_NAME = "warehouse.db"
BACKUP_DIR = "backups"
MAX_BACKUPS = 10
LOG_DIR = "logs"

# --- Логирование (приоритет 1) ---
def _setup_logging():
    os.makedirs(LOG_DIR, exist_ok=True)
    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), LOG_DIR, "app.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stderr),
        ],
    )
    return logging.getLogger("LTO")


logger = None  # инициализируется в main()


def _backup_db(db_path: str) -> None:
    """Автоматическое резервное копирование БД при запуске (приоритет 1)."""
    try:
        base_dir = os.path.dirname(db_path)
        backup_folder = os.path.join(base_dir, BACKUP_DIR)
        os.makedirs(backup_folder, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name = os.path.basename(db_path)
        stem, ext = os.path.splitext(name)
        backup_name = f"{stem}_{stamp}{ext}"
        dest = os.path.join(backup_folder, backup_name)
        shutil.copy2(db_path, dest)
        if logger:
            logger.info("Backup created: %s", dest)
        # Оставляем только последние MAX_BACKUPS
        files = sorted(
            (f for f in os.listdir(backup_folder) if f.startswith(stem) and f.endswith(ext)),
            key=lambda f: os.path.getmtime(os.path.join(backup_folder, f)),
            reverse=True,
        )
        for f in files[MAX_BACKUPS:]:
            try:
                os.remove(os.path.join(backup_folder, f))
            except OSError:
                pass
    except Exception as e:
        if logger:
            logger.exception("Backup failed: %s", e)
DEBUG_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "debug-2a8cbc.log")
DEBUG_SESSION_ID = "2a8cbc"


def _agent_debug_log(hypothesis_id: str, message: str, data: dict | None = None, run_id: str = "run1") -> None:
    """
    Лёгкий логгер для режима отладки: пишет одну строку NDJSON в debug-2a8cbc.log.
    Не бросает исключений при ошибках записи.
    """
    try:
        ts_ms = int(time.time() * 1000)
        payload = {
            "sessionId": DEBUG_SESSION_ID,
            "id": f"log_{ts_ms}",
            "timestamp": ts_ms,
            "location": "main.py",
            "message": message,
            "data": data or {},
            "runId": run_id,
            "hypothesisId": hypothesis_id,
        }
        with open(DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception as e:
        if logger:
            logger.debug("Debug log write failed: %s", e)


def _patch_calendar_arrows(date_edit) -> None:
    """Убирает встроенные чёрные стрелки QCalendarWidget и ставит белые SVG-иконки."""
    cal = date_edit.calendarWidget()
    for name, svg in [
        ("qt_calendar_prevmonth", "arrow-left-white.svg"),
        ("qt_calendar_nextmonth", "arrow-right-white.svg"),
    ]:
        btn = cal.findChild(QToolButton, name)
        if btn:
            btn.setArrowType(Qt.ArrowType.NoArrow)
            # Загружаем SVG, принудительно окрашиваем в белый через CompositionMode
            raw = _icon_from_file(svg, 14).pixmap(QSize(14, 18))
            white = QPixmap(raw.size())
            white.fill(QColor(0, 0, 0, 0))
            p = QPainter(white)
            p.drawPixmap(0, 0, raw)
            p.setCompositionMode(QPainter.CompositionMode.CompositionMode_SourceIn)
            p.fillRect(white.rect(), QColor("#FFFFFF"))
            p.end()
            btn.setIcon(QIcon(white))
            btn.setIconSize(QSize(14, 18))


def _icon_from_file(filename: str, size: int = 20) -> QIcon:
    """Загружает иконку из папки icons/ рядом с main.py. Возвращает пустой QIcon если файл не найден."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(base_dir, "icons", filename)
    if not os.path.exists(path):
        return QIcon()
    pix = QPixmap(path)
    if not pix.isNull() and (pix.width() != size or pix.height() != size):
        pix = pix.scaled(size, size, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
    return QIcon(pix)


class DatabaseManager:
    def __init__(self, path: str):
        self.path = path
        try:
            self.conn = sqlite3.connect(self.path)
            self.conn.row_factory = sqlite3.Row
            self._init_db()
        except Exception as e:
            if logger:
                logger.exception("Database init failed: %s", e)
            raise

    def _init_db(self):
        cur = self.conn.cursor()

        # Подразделения
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS units (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
            """
        )

        # Категории имущества
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
            """
        )
        for cat in ("ЛТО", "ПДИ", "Высотное снаряжение", "Чехлы и палатки"):
            cur.execute("INSERT OR IGNORE INTO categories(name) VALUES (?)", (cat,))

        # Изделия
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                base_code TEXT NOT NULL,
                uom TEXT NOT NULL,
                type TEXT NOT NULL CHECK (type IN ('qty', 'serial'))
            )
            """
        )

        # Варианты (размеры)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS variants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_id INTEGER NOT NULL REFERENCES items(id) ON DELETE CASCADE,
                size_name TEXT NOT NULL,
                full_code TEXT NOT NULL UNIQUE
            )
            """
        )

        # Количественный учет (материальные средства)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS stock_qty (
                variant_id INTEGER PRIMARY KEY REFERENCES variants(id) ON DELETE CASCADE,
                quantity INTEGER NOT NULL DEFAULT 0
            )
            """
        )

        # Серийный учет (основные средства)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS stock_serial (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                variant_id INTEGER NOT NULL REFERENCES variants(id) ON DELETE CASCADE,
                factory_sn TEXT NOT NULL UNIQUE
            )
            """
        )

        # Журнал операций
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS journal (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                op_type TEXT NOT NULL CHECK (op_type IN ('IN','OUT')),
                variant_id INTEGER NOT NULL REFERENCES variants(id),
                quantity INTEGER,
                factory_sn TEXT,
                unit_id INTEGER REFERENCES units(id),
                doc_name TEXT
            )
            """
        )

        # Audit log для критичных операций (приоритет 2)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                entity_id INTEGER,
                details TEXT
            )
            """
        )

        # Заполним пару подразделений по умолчанию
        cur.execute("INSERT OR IGNORE INTO units(name) VALUES (?)", ("Склад-основной",))
        cur.execute("INSERT OR IGNORE INTO units(name) VALUES (?)", ("Цех-1",))

        self.conn.commit()

        # Миграция: добавляем поле doc_name в journal, если его нет (старые БД)
        cur.execute("PRAGMA table_info(journal)")
        cols = [r[1] for r in cur.fetchall()]
        if "doc_name" not in cols:
            cur.execute("ALTER TABLE journal ADD COLUMN doc_name TEXT")
            self.conn.commit()

        # Миграция: добавляем поле category_id в items, если его нет (старые БД)
        cur.execute("PRAGMA table_info(items)")
        item_cols = [r[1] for r in cur.fetchall()]
        if "category_id" not in item_cols:
            cur.execute("ALTER TABLE items ADD COLUMN category_id INTEGER REFERENCES categories(id)")
            self.conn.commit()

        # region agent log
        _agent_debug_log(
            hypothesis_id="H3",
            message="Database initialized",
            data={"db_path": self.path},
        )
        # endregion

        # Если справочник пустой – заполняем образцами
        cur.execute("SELECT COUNT(*) AS cnt FROM items")
        row = cur.fetchone()
        if row and row["cnt"] == 0:
            self._seed_sample_data()

    def _seed_sample_data(self):
        cur = self.conn.cursor()

        cur.execute("SELECT id, name FROM categories")
        cat_by_name = {r["name"]: r["id"] for r in cur.fetchall()}

        # Вспомогательная функция: вставка изделия + вариантов
        def add_item(name, base_code, uom, itype, cat_name, variants):
            """variants: list of (size_name, full_code_10digits)"""
            cat_id = cat_by_name.get(cat_name)
            cur.execute(
                "INSERT INTO items(name, base_code, uom, type, category_id) VALUES (?,?,?,?,?)",
                (name, base_code, uom, itype, cat_id),
            )
            iid = cur.lastrowid
            for size_name, full_code in variants:
                cur.execute(
                    "INSERT INTO variants(item_id, size_name, full_code) VALUES (?,?,?)",
                    (iid, size_name, full_code),
                )
                vid = cur.lastrowid
                if itype == "qty":
                    cur.execute("INSERT INTO stock_qty(variant_id, quantity) VALUES (?,0)", (vid,))

        # ── ЛТО — одежда (размерный ряд, qty) ───────────────────────────────
        # Базовый Н/Н — 10 цифр. Полный Н/Н — 10 цифр (уникальный для каждого размера).
        # Схема полного кода: первые 8 цифр = базовый[0:8], 9-10 = индекс размера (01..08)
        SIZES_C = ["44-170", "46-170", "48-176", "50-176", "52-182", "54-182", "56-188", "58-188"]
        SIZES_F = ["39", "40", "41", "42", "43", "44", "45", "46"]

        def variants_sized(base10, sizes):
            return [(s, base10[:8] + f"{i+1:02d}") for i, s in enumerate(sizes)]

        lto_clothing = [
            ("Костюм лётный летний",             "1776000101", "шт", SIZES_C),
            ("Костюм лётный зимний",             "1776000201", "шт", SIZES_C),
            ("Куртка лётная демисезонная",        "1776000301", "шт", SIZES_C),
            ("Куртка лётная зимняя утеплённая",   "1776000401", "шт", SIZES_C),
            ("Брюки лётные летние",               "1776000501", "шт", SIZES_C),
            ("Брюки лётные зимние",               "1776000601", "шт", SIZES_C),
            ("Комбинезон лётный лёгкий",          "1776000701", "шт", SIZES_C),
            ("Комбинезон технический",            "1776000801", "шт", SIZES_C),
            ("Рубашка форменная длинный рукав",   "1776000901", "шт", SIZES_C),
            ("Рубашка форменная короткий рукав",  "1776001001", "шт", SIZES_C),
            ("Майка форменная",                   "1776001101", "шт", SIZES_C),
            ("Кепи форменное летнее",             "1776001201", "шт", SIZES_C),
            ("Кепи форменное зимнее",             "1776001301", "шт", SIZES_C),
            ("Берет форменный",                   "1776001401", "шт", SIZES_C),
            ("Пилотка",                           "1776001501", "шт", SIZES_C),
            ("Шапка-ушанка меховая",              "1776001601", "шт", SIZES_C),
            ("Перчатки лётные кожаные",           "1776001701", "шт", SIZES_C),
            ("Перчатки зимние утеплённые",        "1776001801", "шт", SIZES_C),
            ("Шарф форменный шерстяной",          "1776001901", "шт", SIZES_C),
            ("Галстук форменный",                 "1776002001", "шт", SIZES_C),
            ("Ремень поясной форменный",          "1776002101", "шт", SIZES_C),
            ("Носки форменные хлопок",            "1776002201", "шт", SIZES_C),
            ("Носки форменные шерстяные",         "1776002301", "шт", SIZES_C),
            ("Нательное бельё комплект летний",   "1776002401", "шт", SIZES_C),
            ("Нательное бельё комплект зимний",   "1776002501", "шт", SIZES_C),
        ]
        lto_footwear = [
            ("Ботинки лётные хромовые",   "1776100101", "пар", SIZES_F),
            ("Сапоги форменные хромовые", "1776100201", "пар", SIZES_F),
            ("Полуботинки форменные",     "1776100301", "пар", SIZES_F),
            ("Ботинки зимние утеплённые", "1776100401", "пар", SIZES_F),
        ]

        for name, base10, uom, sizes in lto_clothing:
            add_item(name, base10, uom, "qty", "ЛТО", variants_sized(base10, sizes))
        for name, base10, uom, sizes in lto_footwear:
            add_item(name, base10, uom, "qty", "ЛТО", variants_sized(base10, sizes))

        # ── ПДИ — серийный учёт ──────────────────────────────────────────────
        pdi_items = [
            ("Парашют десантный Д-10",           "9306000101"),
            ("Парашют десантный Д-6 серия 4",    "9306000201"),
            ("Парашют запасной З-5",             "9306000301"),
            ("Парашют запасной З-6П",            "9306000401"),
            ("Страховочный прибор ППК-У",        "9306000501"),
            ("Страховочный прибор ППКУ-165А",    "9306000601"),
            ("Ранец парашюта основной",          "9306000701"),
            ("Ранец парашюта запасной",          "9306000801"),
            ("Карабин страховочный десантный",   "9306000901"),
            ("Грузовой контейнер УГКП-500",      "9306001001"),
            ("Подвесная система ПС-Д10",         "9306001101"),
            ("Вытяжной фал 15м",                 "9306001201"),
            ("Кольцо вытяжного троса",           "9306001301"),
            ("Шпилька предохранительная",        "9306001401"),
            ("Соединительное звено",             "9306001501"),
            ("Слинг стропы 7мм",                 "9306001601"),
            ("Чехол стабилизирующего купола",    "9306001701"),
            ("Укладочная доска",                 "9306001801"),
            ("Прибор контроля укладки",          "9306001901"),
            ("Сумка укладчика",                  "9306002001"),
            ("Журнал контроля парашюта",         "9306002101"),
            ("Паспорт парашюта",                 "9306002201"),
        ]
        for name, base10 in pdi_items:
            add_item(name, base10, "шт", "serial", "ПДИ", [("UNI", base10)])

        # ── Высотное снаряжение ──────────────────────────────────────────────
        high_serial = [
            ("Каска альпинистская PETZL Vertex",            "6506000101"),
            ("Каска защитная CAMP Ares",                    "6506000201"),
            ("Беседка страховочная Singing Rock",           "6506000301"),
            ("Беседка комбинированная CAMP Jasper CR3",     "6506000401"),
            ("Жумар левый PETZL Ascension",                 "6506000501"),
            ("Жумар правый PETZL Ascension",                "6506000601"),
            ("Рюкзак-контейнер для верёвки 70л",            "6506000701"),
            ("Спусковое устройство PETZL Pirana",           "6506000801"),
            ("Спусковое устройство CAMP Raft",              "6506000901"),
            ("Тянущее устройство PETZL BASIC",              "6506001001"),
            ("Блок-ролик одинарный PETZL Fixe",             "6506001101"),
            ("Блок-ролик двойной PETZL Tandem",             "6506001201"),
            ("Страховочное устройство PETZL GriGri",        "6506001301"),
            ("Карабин D-образный муфтованный",              "6506001401"),
            ("Карабин HMS муфтованный",                     "6506001501"),
        ]
        high_qty = [
            ("Верёвка статическая 10мм (бухта 50м)",  "5607000101", "м"),
            ("Верёвка статическая 11мм (бухта 50м)",  "5607000201", "м"),
            ("Верёвка динамическая 9мм (бухта 60м)",  "5607000301", "м"),
            ("Стропа петлевая 16мм (бухта 50м)",      "5607000401", "м"),
            ("Репшнур 6мм (бухта 30м)",               "5607000501", "м"),
            ("Петля готовая из стропы 120см",         "5607000601", "шт"),
            ("Лента стропа 25мм (бухта 50м)",         "5607000701", "м"),
        ]
        for name, base10 in high_serial:
            add_item(name, base10, "шт", "serial", "Высотное снаряжение", [("UNI", base10)])
        for name, base10, uom in high_qty:
            add_item(name, base10, uom, "qty", "Высотное снаряжение", [("UNI", base10)])

        # ── Чехлы и палатки — qty ────────────────────────────────────────────
        tents_items = [
            ("Палатка армейская 2-местная УСБ-56",   "6301000101", "шт"),
            ("Палатка армейская 4-местная ПБ-1",     "6301000201", "шт"),
            ("Палатка 10-местная ПУС-2",             "6301000301", "шт"),
            ("Палатка командно-штабная КШМ-4",       "6301000401", "шт"),
            ("Тент армейский маскировочный 12×12м",  "6301000501", "шт"),
            ("Тент армейский защитный 6×9м",         "6301000601", "шт"),
            ("Чехол для парашюта Д-10",              "6301000701", "шт"),
            ("Чехол для парашюта З-5",               "6301000801", "шт"),
            ("Чехол для карабина АКМ",               "6301000901", "шт"),
            ("Чехол для карабина АК-74",             "6301001001", "шт"),
            ("Чехол для пулемёта РПК",               "6301001101", "шт"),
            ("Чехол для снайперской винтовки",       "6301001201", "шт"),
            ("Чехол для бинокля",                    "6301001301", "шт"),
            ("Чехол для радиостанции Р-168",         "6301001401", "шт"),
            ("Мешок вещевой армейский",              "6301001501", "шт"),
            ("Рюкзак полевой 60л",                   "6301001601", "шт"),
            ("Рюкзак десантный РД-54",               "6301001701", "шт"),
            ("Сумка полевая командира",              "6301001801", "шт"),
            ("Сумка медицинская войсковая",          "6301001901", "шт"),
            ("Спальный мешок летний",                "6301002001", "шт"),
            ("Спальный мешок зимний -20°C",          "6301002101", "шт"),
            ("Коврик пенополиуретановый",            "6301002201", "шт"),
            ("Накидка плащ-палатка",                 "6301002301", "шт"),
            ("Куртка дождевая",                      "6301002401", "шт"),
            ("Чехол для касок комплект 10шт",        "6301002501", "компл"),
        ]
        for name, base10, uom in tents_items:
            add_item(name, base10, uom, "qty", "Чехлы и палатки", [("UNI", base10)])

        self.conn.commit()

    # --- Items & Variants ---

    def get_categories(self):
        cur = self.conn.cursor()
        cur.execute("SELECT id, name FROM categories ORDER BY id")
        return cur.fetchall()

    def get_items(self):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT i.id, i.name, i.base_code, i.uom, i.type,
                   COALESCE(c.name, '—') AS category_name, i.category_id
            FROM items i
            LEFT JOIN categories c ON c.id = i.category_id
            ORDER BY c.name COLLATE NOCASE, i.name COLLATE NOCASE
            """
        )
        return cur.fetchall()

    def add_item(self, name: str, base_code: str, uom: str, item_type: str, category_id: int | None = None):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO items(name, base_code, uom, type, category_id) VALUES (?, ?, ?, ?, ?)",
            (name.strip(), base_code.strip(), uom.strip(), item_type, category_id),
        )
        iid = cur.lastrowid
        self.conn.commit()
        self._audit("ITEM_INSERT", "items", iid, f"name={name.strip()} base_code={base_code.strip()}")
        return iid

    def get_variants_for_item(self, item_id: int):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT id, size_name, full_code
            FROM variants
            WHERE item_id = ?
            ORDER BY size_name COLLATE NOCASE
            """,
            (item_id,),
        )
        return cur.fetchall()

    def get_item(self, item_id: int):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM items WHERE id = ?", (item_id,))
        return cur.fetchone()

    def add_variant(self, item_id: int, size_name: str, full_code: str, item_type: str):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO variants(item_id, size_name, full_code) VALUES (?, ?, ?)",
            (item_id, size_name.strip(), full_code.strip()),
        )
        variant_id = cur.lastrowid
        # Для количественного учета сразу создаем запись остатков (0)
        if item_type == "qty":
            cur.execute(
                "INSERT OR IGNORE INTO stock_qty(variant_id, quantity) VALUES (?, 0)",
                (variant_id,),
            )
        self.conn.commit()
        self._audit("VARIANT_INSERT", "variants", variant_id, f"item_id={item_id} size={size_name.strip()} full_code={full_code.strip()}")
        return variant_id

    # --- Units ---

    def get_units(self):
        cur = self.conn.cursor()
        cur.execute("SELECT id, name FROM units ORDER BY name COLLATE NOCASE")
        return cur.fetchall()

    def add_unit(self, name: str):
        cur = self.conn.cursor()
        cur.execute("INSERT INTO units(name) VALUES (?)", (name.strip(),))
        uid = cur.lastrowid
        self.conn.commit()
        self._audit("UNIT_INSERT", "units", uid, f"name={name.strip()}")
        return uid

    def delete_unit(self, unit_id: int):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM units WHERE id = ?", (unit_id,))
        self.conn.commit()
        self._audit("UNIT_DELETE", "units", unit_id, "")

    def has_journal_entries_for_item(self, item_id: int) -> bool:
        """Проверяет, есть ли в журнале операции по любому варианту изделия."""
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT COUNT(*) AS cnt FROM journal j
            JOIN variants v ON v.id = j.variant_id
            WHERE v.item_id = ?
            """,
            (item_id,),
        )
        return (cur.fetchone()["cnt"] or 0) > 0

    def has_journal_entries_for_variant(self, variant_id: int) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT COUNT(*) AS cnt FROM journal WHERE variant_id = ?",
            (variant_id,),
        )
        return (cur.fetchone()["cnt"] or 0) > 0

    def delete_item(self, item_id: int):
        """Удаляет изделие и все связанные варианты, остатки и записи журнала."""
        cur = self.conn.cursor()
        cur.execute("SELECT id FROM variants WHERE item_id = ?", (item_id,))
        variant_ids = [r["id"] for r in cur.fetchall()]
        for vid in variant_ids:
            cur.execute("DELETE FROM journal WHERE variant_id = ?", (vid,))
            cur.execute("DELETE FROM stock_qty WHERE variant_id = ?", (vid,))
            cur.execute("DELETE FROM stock_serial WHERE variant_id = ?", (vid,))
        cur.execute("DELETE FROM variants WHERE item_id = ?", (item_id,))
        cur.execute("DELETE FROM items WHERE id = ?", (item_id,))
        self.conn.commit()
        self._audit("ITEM_DELETE", "items", item_id, f"variants={variant_ids}")

    def delete_variant(self, variant_id: int):
        """Удаляет вариант (размер) и все связанные остатки и записи журнала."""
        cur = self.conn.cursor()
        cur.execute("DELETE FROM journal WHERE variant_id = ?", (variant_id,))
        cur.execute("DELETE FROM stock_qty WHERE variant_id = ?", (variant_id,))
        cur.execute("DELETE FROM stock_serial WHERE variant_id = ?", (variant_id,))
        cur.execute("DELETE FROM variants WHERE id = ?", (variant_id,))
        self.conn.commit()
        self._audit("VARIANT_DELETE", "variants", variant_id, "")

    def get_or_create_unit(self, name: str | None):
        if not name:
            return None
        name = name.strip()
        if not name:
            return None
        cur = self.conn.cursor()
        cur.execute("SELECT id FROM units WHERE name = ?", (name,))
        row = cur.fetchone()
        if row:
            return row["id"]
        cur.execute("INSERT INTO units(name) VALUES (?)", (name,))
        self.conn.commit()
        return cur.lastrowid

    # --- Search & Info ---

    def search_variants(self, text: str, only_in_stock: bool = False):
        pattern = f"%{text.strip()}%"
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                v.id AS variant_id,
                v.full_code,
                v.size_name,
                i.name AS item_name,
                i.type AS item_type,
                i.uom AS uom,
                COALESCE(c.name, '—') AS category_name,
                CASE
                    WHEN i.type = 'qty' THEN COALESCE(sq.quantity, 0)
                    ELSE (SELECT COUNT(*) FROM stock_serial ss WHERE ss.variant_id = v.id)
                END AS stock_value
            FROM variants v
            JOIN items i ON v.item_id = i.id
            LEFT JOIN categories c ON c.id = i.category_id
            LEFT JOIN stock_qty sq ON sq.variant_id = v.id
            WHERE v.full_code LIKE ? OR i.name LIKE ?
            ORDER BY i.name COLLATE NOCASE, v.size_name COLLATE NOCASE
            """,
            (pattern, pattern),
        )
        rows = cur.fetchall()
        if only_in_stock:
            rows = [r for r in rows if (r["stock_value"] or 0) > 0]
        return rows

    def get_variant_with_item(self, variant_id: int):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                v.id AS variant_id,
                v.full_code,
                v.size_name,
                i.id AS item_id,
                i.name AS item_name,
                i.base_code,
                i.uom,
                i.type AS item_type
            FROM variants v
            JOIN items i ON v.item_id = i.id
            WHERE v.id = ?
            """,
            (variant_id,),
        )
        return cur.fetchone()

    # --- Stock operations ---

    def get_qty_stock(self, variant_id: int) -> int:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT quantity FROM stock_qty WHERE variant_id = ?", (variant_id,)
        )
        row = cur.fetchone()
        return int(row["quantity"]) if row else 0

    def adjust_qty_stock(self, variant_id: int, delta: int):
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO stock_qty(variant_id, quantity)
            VALUES (?, ?)
            ON CONFLICT(variant_id) DO UPDATE SET quantity = quantity + excluded.quantity
            """,
            (variant_id, delta),
        )
        self.conn.commit()

    def serial_exists(self, factory_sn: str) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id FROM stock_serial WHERE factory_sn = ?", (factory_sn.strip(),)
        )
        return cur.fetchone() is not None

    def serial_exists_for_variant(self, variant_id: int, factory_sn: str) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id FROM stock_serial WHERE variant_id = ? AND factory_sn = ?",
            (variant_id, factory_sn.strip()),
        )
        return cur.fetchone() is not None

    def add_serial(self, variant_id: int, factory_sn: str):
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO stock_serial(variant_id, factory_sn) VALUES (?, ?)",
            (variant_id, factory_sn.strip()),
        )
        self.conn.commit()

    def remove_serial(self, variant_id: int, factory_sn: str) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            "DELETE FROM stock_serial WHERE variant_id = ? AND factory_sn = ?",
            (variant_id, factory_sn.strip()),
        )
        deleted = cur.rowcount > 0
        self.conn.commit()
        return deleted

    # --- Journal ---

    def _audit(self, action: str, entity_type: str, entity_id: int | None, details: str = ""):
        try:
            cur = self.conn.cursor()
            cur.execute(
                """
                INSERT INTO audit_log(ts, action, entity_type, entity_id, details)
                VALUES (?, ?, ?, ?, ?)
                """,
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), action, entity_type, entity_id or 0, details),
            )
            self.conn.commit()
        except Exception as e:
            if logger:
                logger.exception("Audit log write failed: %s", e)

    def add_journal_record(
        self,
        op_type: str,
        variant_id: int,
        quantity: int | None,
        factory_sn: str | None,
        unit_id: int | None,
        doc_name: str,
    ):
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO journal(date, op_type, variant_id, quantity, factory_sn, unit_id, doc_name)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                op_type,
                variant_id,
                quantity,
                factory_sn.strip() if factory_sn else None,
                unit_id,
                doc_name.strip(),
            ),
        )
        jid = cur.lastrowid
        self.conn.commit()
        self._audit("JOURNAL_INSERT", "journal", jid, f"op={op_type} variant_id={variant_id} doc={doc_name.strip()}")

    # --- Stock view ---

    def get_stock_view(self):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                i.id AS item_id,
                i.name AS item_name,
                i.base_code,
                i.type AS item_type,
                i.uom AS uom,
                COALESCE(c.name, '—') AS category_name,
                COALESCE(
                    CASE
                        WHEN i.type = 'qty' THEN SUM(sq.quantity)
                        ELSE COUNT(ss.id)
                    END,
                    0
                ) AS stock_value
            FROM items i
            LEFT JOIN categories c ON c.id = i.category_id
            LEFT JOIN variants v ON v.item_id = i.id
            LEFT JOIN stock_qty sq ON sq.variant_id = v.id
            LEFT JOIN stock_serial ss ON ss.variant_id = v.id
            GROUP BY i.id, i.name, i.base_code, i.type, i.uom, c.name
            HAVING stock_value > 0
            ORDER BY c.name COLLATE NOCASE, i.name COLLATE NOCASE
            """
        )
        return cur.fetchall()

    def get_journal_view(
        self,
        limit: int = 200,
        date_from: str | None = None,
        date_to: str | None = None,
        unit_id: int | None = None,
    ):
        cur = self.conn.cursor()
        where_parts = []
        params = []
        if date_from:
            where_parts.append("date(j.date) >= ?")
            params.append(date_from)
        if date_to:
            where_parts.append("date(j.date) <= ?")
            params.append(date_to)
        if unit_id is not None:
            where_parts.append("j.unit_id = ?")
            params.append(unit_id)
        where_sql = " WHERE " + " AND ".join(where_parts) if where_parts else ""
        params.append(limit)
        cur.execute(
            f"""
            SELECT
                j.id AS id,
                j.date AS date,
                j.op_type AS op_type,
                j.doc_name AS doc_name,
                v.size_name AS size_name,
                v.full_code AS full_code,
                i.name AS item_name,
                i.type AS item_type,
                j.quantity AS quantity,
                j.factory_sn AS factory_sn,
                u.name AS unit_name
            FROM journal j
            JOIN variants v ON j.variant_id = v.id
            JOIN items i ON v.item_id = i.id
            LEFT JOIN units u ON j.unit_id = u.id
            {where_sql}
            ORDER BY j.date DESC, j.id DESC
            LIMIT ?
            """,
            tuple(params),
        )
        return cur.fetchall()

    def get_serials_for_variant(self, variant_id: int):
        """Возвращает все серийные номера на складе для данного варианта."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT factory_sn FROM stock_serial WHERE variant_id = ? ORDER BY factory_sn COLLATE NOCASE",
            (variant_id,),
        )
        return cur.fetchall()

    def get_serials_for_item(self, item_id: int):
        """Возвращает все S/N на складе по всем вариантам изделия (с размером и полным кодом)."""
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT ss.factory_sn, v.size_name, v.full_code
            FROM stock_serial ss
            JOIN variants v ON v.id = ss.variant_id
            WHERE v.item_id = ?
            ORDER BY v.size_name COLLATE NOCASE, ss.factory_sn COLLATE NOCASE
            """,
            (item_id,),
        )
        return cur.fetchall()

    def get_item_stock_details(self, item_id: int):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                v.id AS variant_id,
                v.full_code,
                v.size_name,
                i.uom AS uom,
                COALESCE(
                    CASE
                        WHEN i.type = 'qty' THEN sq.quantity
                        ELSE COUNT(ss.id)
                    END,
                    0
                ) AS stock_value
            FROM variants v
            JOIN items i ON v.item_id = i.id
            LEFT JOIN stock_qty sq ON sq.variant_id = v.id
            LEFT JOIN stock_serial ss ON ss.variant_id = v.id
            WHERE v.item_id = ?
            GROUP BY v.id, v.full_code, v.size_name, i.uom, sq.quantity
            ORDER BY v.size_name COLLATE NOCASE
            """,
            (item_id,),
        )
        return cur.fetchall()


# --- Экспорт отчётов (приоритет 2): Excel / PDF ---
def _export_journal_excel(db: DatabaseManager, path: str, date_from: str, date_to: str, unit_id: int | None) -> bool:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        if logger:
            logger.warning("openpyxl not installed: pip install openpyxl")
        return False
    try:
        rows = db.get_journal_view(limit=10000, date_from=date_from, date_to=date_to, unit_id=unit_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Журнал операций"
        headers = ["Дата", "Операция", "Документ", "Название", "Размер", "Кол-во", "Подразделение"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
            ws.cell(1, c).font = Font(bold=True)
        for r, row in enumerate(rows, 2):
            op_text = "Приход" if row["op_type"] == "IN" else "Выдача"
            qty = row["quantity"] if row["quantity"] is not None else (1 if row["factory_sn"] else 0)
            ws.cell(r, 1, row["date"])
            ws.cell(r, 2, op_text)
            ws.cell(r, 3, row["doc_name"] or "")
            ws.cell(r, 4, row["item_name"])
            ws.cell(r, 5, row["size_name"])
            ws.cell(r, 6, qty)
            ws.cell(r, 7, row["unit_name"] or "")
        wb.save(path)
        if logger:
            logger.info("Exported journal to Excel: %s", path)
        return True
    except Exception as e:
        if logger:
            logger.exception("Export journal Excel failed: %s", e)
        return False


def _register_pdf_font() -> str:
    """
    Регистрирует Arial из системных шрифтов Windows для поддержки кириллицы.
    Возвращает имя зарегистрированного шрифта.
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "ArialUnicode"
    try:
        pdfmetrics.getFont(font_name)
        return font_name
    except Exception:
        pass

    candidates = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/Arial.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont(font_name, path))
            return font_name

    return "Helvetica"


def _export_journal_pdf(db: DatabaseManager, path: str, date_from: str, date_to: str, unit_id: int | None) -> bool:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
    except ImportError:
        if logger:
            logger.warning("reportlab not installed: pip install reportlab")
        return False
    try:
        font = _register_pdf_font()
        rows = db.get_journal_view(limit=10000, date_from=date_from, date_to=date_to, unit_id=unit_id)

        doc = SimpleDocTemplate(
            path,
            pagesize=landscape(A4),
            leftMargin=10 * mm, rightMargin=10 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
        )

        title_style = ParagraphStyle("title", fontName=font, fontSize=12, spaceAfter=6)
        header = Paragraph(
            f"Журнал операций ({date_from} — {date_to})",
            title_style,
        )

        col_headers = ["Дата", "Операция", "Документ", "Название", "Размер", "Кол-во", "Подразделение"]
        data = [col_headers]
        for row in rows:
            op_text = "Приход" if row["op_type"] == "IN" else "Выдача"
            qty = row["quantity"] if row["quantity"] is not None else (1 if row["factory_sn"] else 0)
            data.append([
                row["date"], op_text, row["doc_name"] or "", row["item_name"],
                row["size_name"], str(qty), row["unit_name"] or "",
            ])

        col_widths = [35*mm, 22*mm, 35*mm, 70*mm, 22*mm, 18*mm, 40*mm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f7aec")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), font),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4ff")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d5e8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ]))

        doc.build([header, Spacer(1, 4*mm), t])
        if logger:
            logger.info("Exported journal to PDF: %s", path)
        return True
    except Exception as e:
        if logger:
            logger.exception("Export journal PDF failed: %s", e)
        return False


def _export_stock_excel(db: DatabaseManager, path: str) -> bool:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        if logger:
            logger.warning("openpyxl not installed: pip install openpyxl")
        return False
    try:
        rows = db.get_stock_view()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Остатки"
        headers = ["Н/Н (базовый)", "Название", "Остаток", "Ед. изм."]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
            ws.cell(1, c).font = Font(bold=True)
        for r, row in enumerate(rows, 2):
            ws.cell(r, 1, row["base_code"])
            ws.cell(r, 2, row["item_name"])
            ws.cell(r, 3, row["stock_value"])
            ws.cell(r, 4, row["uom"])
        wb.save(path)
        if logger:
            logger.info("Exported stock to Excel: %s", path)
        return True
    except Exception as e:
        if logger:
            logger.exception("Export stock Excel failed: %s", e)
        return False


def _export_stock_pdf(db: DatabaseManager, path: str) -> bool:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
    except ImportError:
        if logger:
            logger.warning("reportlab not installed: pip install reportlab")
        return False
    try:
        font = _register_pdf_font()

        doc = SimpleDocTemplate(
            path,
            pagesize=A4,
            leftMargin=10 * mm, rightMargin=10 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
        )

        title_style = ParagraphStyle("title", fontName=font, fontSize=12, spaceAfter=6)
        sub_style = ParagraphStyle("sub", fontName=font, fontSize=8, textColor=colors.HexColor("#555555"), spaceAfter=2)

        header = Paragraph("Остатки на складе", title_style)
        header_date = Paragraph(
            f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            sub_style,
        )

        base_rows = db.get_stock_view()

        elements = [header, header_date, Spacer(1, 4 * mm)]

        # Одинаковые ширины колонок для заголовка и размерного ряда
        COL_W = [35*mm, 70*mm, 30*mm, 20*mm]  # итого 155 mm

        for row in base_rows:
            details = db.get_item_stock_details(row["item_id"])

            # Заголовок изделия — те же 4 колонки, первые 2 объединены через SPAN
            item_data = [[
                row["base_code"],
                row["item_name"],
                f"Итого: {row['stock_value']}",
                row["uom"],
            ]]
            item_table = Table(item_data, colWidths=COL_W)
            item_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1f7aec")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#1764c0")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(item_table)

            # Размерный ряд — те же 4 колонки, те же ширины
            if details:
                size_data = [["  Размер", "Н/Н (полный)", "Остаток", "Ед. изм."]]
                for d in details:
                    size_data.append([
                        f"  {d['size_name']}",
                        d["full_code"],
                        str(d["stock_value"]),
                        d["uom"],
                    ])
                size_table = Table(size_data, colWidths=COL_W)
                size_table.setStyle(TableStyle([
                    ("FONTNAME", (0, 0), (-1, -1), font),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8edf7")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7ff")]),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d0d5e8")),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ]))
                elements.append(size_table)

            elements.append(Spacer(1, 3 * mm))

        doc.build(elements)
        if logger:
            logger.info("Exported stock to PDF: %s", path)
        return True
    except Exception as e:
        if logger:
            logger.exception("Export stock PDF failed: %s", e)
        return False


class NewItemDialog(QDialog):
    def __init__(self, db: "DatabaseManager", parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Создать изделие")
        self.item_type = "qty"
        self._categories = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        form = QFormLayout()
        self.name_edit = QLineEdit()

        self.base_code_edit = QLineEdit()
        self.base_code_edit.setMaxLength(10)
        self.base_code_edit.setPlaceholderText("10 цифр, например 1776184605")
        self.base_code_edit.setInputMask("9999999999")

        self.uom_edit = QLineEdit()
        self.uom_edit.setText("шт")

        form.addRow("Название:", self.name_edit)
        form.addRow("Н/Н (базовый):", self.base_code_edit)
        form.addRow("Ед. изм.:", self.uom_edit)

        # Выбор категории
        self.category_combo = QComboBox()
        self._categories = self.db.get_categories()
        self.category_combo.addItem("— без категории —", None)
        for cat in self._categories:
            self.category_combo.addItem(cat["name"], cat["id"])
        form.addRow("Категория:", self.category_combo)

        type_layout = QHBoxLayout()
        self.qty_radio = QRadioButton("Мат. средства")
        self.serial_radio = QRadioButton("Основные средства")
        self.qty_radio.setChecked(True)
        type_layout.addWidget(self.qty_radio)
        type_layout.addWidget(self.serial_radio)

        type_group = QButtonGroup(self)
        type_group.addButton(self.qty_radio)
        type_group.addButton(self.serial_radio)

        type_wrapper = QWidget()
        type_wrapper.setLayout(type_layout)
        form.addRow("Тип учета:", type_wrapper)

        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        self.ok_btn = QPushButton("OK")
        self.cancel_btn = QPushButton("Отмена")
        btn_layout.addWidget(self.ok_btn)
        btn_layout.addWidget(self.cancel_btn)
        layout.addLayout(btn_layout)

        self.ok_btn.clicked.connect(self.accept)
        self.cancel_btn.clicked.connect(self.reject)

    def get_data(self):
        name = self.name_edit.text().strip()
        base_code = self.base_code_edit.text().strip()
        uom = self.uom_edit.text().strip()
        item_type = "qty" if self.qty_radio.isChecked() else "serial"
        category_id = self.category_combo.currentData()
        return name, base_code, uom, item_type, category_id

    def accept(self):
        name, base_code, uom, _, _cat = self.get_data()
        if not name or not uom:
            QMessageBox.warning(self, "Ошибка", "Заполните все поля.")
            return
        if not base_code.isdigit() or len(base_code) != 10:
            QMessageBox.warning(self, "Ошибка", "Н/Н (базовый) должен содержать ровно 10 цифр.")
            return
        super().accept()


class NewVariantDialog(QDialog):
    def __init__(self, base_code: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить размер")
        self.base_code = base_code
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.size_edit = QLineEdit()
        self.size_edit.setPlaceholderText("например: 44-170 или UNI")

        self.full_code_edit = QLineEdit()
        self.full_code_edit.setMaxLength(10)
        self.full_code_edit.setPlaceholderText("10 цифр, например 1776184606")
        self.full_code_edit.setInputMask("9999999999")

        form.addRow("Размер:", self.size_edit)
        form.addRow("Н/Н (полный):", self.full_code_edit)

        hint = QLabel(f"Базовый Н/Н изделия: <b>{self.base_code}</b>")
        hint.setStyleSheet("color: #6B778C; font-size: 11px;")
        form.addRow("", hint)

        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        self.ok_btn = QPushButton("OK")
        self.cancel_btn = QPushButton("Отмена")
        btn_layout.addWidget(self.ok_btn)
        btn_layout.addWidget(self.cancel_btn)
        layout.addLayout(btn_layout)

        self.ok_btn.clicked.connect(self.accept)
        self.cancel_btn.clicked.connect(self.reject)

    def get_data(self):
        size = self.size_edit.text().strip()
        full_code = self.full_code_edit.text().strip()
        return size, full_code

    def accept(self):
        size, full_code = self.get_data()
        if not size:
            QMessageBox.warning(self, "Ошибка", "Укажите размер.")
            return
        if not full_code.isdigit() or len(full_code) != 10:
            QMessageBox.warning(self, "Ошибка", "Н/Н (полный) должен содержать ровно 10 цифр.")
            return
        super().accept()


class NomenclatureTab(QWidget):
    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self.current_item_id: int | None = None
        self._variants: list = []
        self._build_ui()
        self.load_items()

    def _build_ui(self):
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(24, 20, 24, 20)
        main_layout.setSpacing(16)

        # Левая часть: изделия
        left_layout = QVBoxLayout()
        left_layout.addWidget(QLabel("Изделия:"))
        self.items_list = QListWidget()
        left_layout.addWidget(self.items_list)
        items_btn_row = QHBoxLayout()
        self.new_item_btn = QPushButton("Создать изделие")
        self.delete_item_btn = QPushButton("Удалить")
        self.delete_item_btn.setObjectName("DangerBtn")
        items_btn_row.addWidget(self.new_item_btn)
        items_btn_row.addWidget(self.delete_item_btn)
        left_layout.addLayout(items_btn_row)

        # Правая часть: варианты
        right_layout = QVBoxLayout()
        right_layout.addWidget(QLabel("Варианты (размеры):"))
        self.variants_list = QListWidget()
        right_layout.addWidget(self.variants_list)
        variants_btn_row = QHBoxLayout()
        self.new_variant_btn = QPushButton("Добавить размер")
        self.delete_variant_btn = QPushButton("Удалить")
        self.delete_variant_btn.setObjectName("DangerBtn")
        variants_btn_row.addWidget(self.new_variant_btn)
        variants_btn_row.addWidget(self.delete_variant_btn)
        right_layout.addLayout(variants_btn_row)

        main_layout.addLayout(left_layout, 1)
        main_layout.addLayout(right_layout, 1)

        self.items_list.currentRowChanged.connect(self.on_item_selected)
        self.new_item_btn.clicked.connect(self.on_new_item)
        self.delete_item_btn.clicked.connect(self.on_delete_item)
        self.new_variant_btn.clicked.connect(self.on_new_variant)
        self.delete_variant_btn.clicked.connect(self.on_delete_variant)

    def load_items(self):
        self.items_list.clear()
        self.current_item_id = None
        self._items = self.db.get_items()
        for row in self._items:
            text = f"{row['name']}  [{row['base_code']}]"
            self.items_list.addItem(text)
        if self._items:
            self.items_list.setCurrentRow(0)

    def load_variants(self, item_id: int | None):
        self.variants_list.clear()
        self._variants = []
        if item_id is None:
            return
        self._variants = self.db.get_variants_for_item(item_id)
        for v in self._variants:
            text = f"{v['size_name']}  [{v['full_code']}]"
            self.variants_list.addItem(text)

    def on_item_selected(self, index: int):
        if index < 0 or index >= len(getattr(self, "_items", [])):
            self.current_item_id = None
            self.variants_list.clear()
            return
        item_row = self._items[index]
        self.current_item_id = item_row["id"]
        self.load_variants(self.current_item_id)

    def on_new_item(self):
        dlg = NewItemDialog(self.db, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            name, base_code, uom, item_type, category_id = dlg.get_data()
            try:
                new_id = self.db.add_item(name, base_code, uom, item_type, category_id)
            except sqlite3.IntegrityError as e:
                if logger:
                    logger.warning("Add item failed: %s", e)
                QMessageBox.warning(self, "Ошибка", f"Не удалось создать изделие:\n{e}")
                return
            self.load_items()

            # Сразу предлагаем добавить первый вариант — без варианта изделие
            # не появится в поиске при проведении операций
            reply = QMessageBox.question(
                self, "Добавить вариант",
                f"Изделие «{name}» создано.\n\n"
                "Чтобы проводить операции, нужен хотя бы один вариант (размер).\n"
                "Добавить вариант сейчас?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )
            if reply == QMessageBox.StandardButton.Yes:
                # Выбираем только что созданное изделие и открываем диалог варианта
                for i, r in enumerate(self._items):
                    if r["id"] == new_id:
                        self.items_list.setCurrentRow(i)
                        break
                self.on_new_variant()

    def on_new_variant(self):
        if self.current_item_id is None:
            QMessageBox.warning(self, "Ошибка", "Сначала выберите изделие.")
            return
        item_row = self.db.get_item(self.current_item_id)
        if item_row is None:
            QMessageBox.warning(self, "Ошибка", "Изделие не найдено.")
            return
        base_code = item_row["base_code"]
        item_type = item_row["type"]

        dlg = NewVariantDialog(base_code, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            size_name, full_code = dlg.get_data()
            try:
                self.db.add_variant(self.current_item_id, size_name, full_code, item_type)
            except sqlite3.IntegrityError as e:
                if logger:
                    logger.warning("Add variant duplicate: %s", e)
                QMessageBox.warning(
                    self,
                    "Ошибка",
                    "Вариант с таким полным кодом уже существует.",
                )
                return
            self.load_variants(self.current_item_id)

    def on_delete_item(self):
        idx = self.items_list.currentRow()
        if idx < 0 or idx >= len(getattr(self, "_items", [])):
            QMessageBox.warning(self, "Удаление", "Выберите изделие для удаления.")
            return
        item_row = self._items[idx]
        item_id = item_row["id"]
        item_name = item_row["name"]

        has_history = self.db.has_journal_entries_for_item(item_id)
        if has_history:
            msg = (
                f"Изделие «{item_name}» имеет историю операций.\n"
                "Удаление удалит все варианты, остатки и записи журнала.\n\n"
                "Вы уверены?"
            )
        else:
            msg = f"Удалить изделие «{item_name}» и все его варианты?"

        reply = QMessageBox.question(
            self, "Подтверждение удаления", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            self.db.delete_item(item_id)
            if logger:
                logger.info("Item deleted: id=%s name=%s", item_id, item_name)
        except Exception as e:
            if logger:
                logger.exception("Delete item failed: %s", e)
            QMessageBox.critical(self, "Ошибка", f"Не удалось удалить изделие:\n{e}")
            return

        self.load_items()

    def on_delete_variant(self):
        idx = self.variants_list.currentRow()
        if idx < 0 or idx >= len(self._variants):
            QMessageBox.warning(self, "Удаление", "Выберите размер для удаления.")
            return
        v = self._variants[idx]
        variant_id = v["id"]
        label = f"{v['size_name']} [{v['full_code']}]"

        has_history = self.db.has_journal_entries_for_variant(variant_id)
        if has_history:
            msg = (
                f"Вариант «{label}» имеет историю операций.\n"
                "Удаление удалит остатки и записи журнала по этому варианту.\n\n"
                "Вы уверены?"
            )
        else:
            msg = f"Удалить вариант «{label}»?"

        reply = QMessageBox.question(
            self, "Подтверждение удаления", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            self.db.delete_variant(variant_id)
            if logger:
                logger.info("Variant deleted: id=%s label=%s", variant_id, label)
        except Exception as e:
            if logger:
                logger.exception("Delete variant failed: %s", e)
            QMessageBox.critical(self, "Ошибка", f"Не удалось удалить вариант:\n{e}")
            return

        self.load_variants(self.current_item_id)


class OperationDetailDialog(QDialog):
    """Модальное окно с деталями одной операции."""

    def __init__(self, op_data: dict, parent=None):
        super().__init__(parent)
        op_type  = "ПРИХОД" if op_data["op_type"] == "IN" else "ВЫДАЧА"
        self.setWindowTitle(f"Документ № {op_data['doc_name']}")
        self.setMinimumWidth(640)
        self.setMinimumHeight(400)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(16)

        # ── Шапка документа ──────────────────────────────────────────────────
        header_frame = QFrame()
        header_frame.setStyleSheet(
            "QFrame { background:#F4F5F7; border:1px solid #DFE1E6; border-radius:4px; }"
        )
        hf_layout = QGridLayout(header_frame)
        hf_layout.setContentsMargins(16, 12, 16, 12)
        hf_layout.setHorizontalSpacing(32)
        hf_layout.setVerticalSpacing(6)

        def _lbl_key(text):
            l = QLabel(text)
            l.setStyleSheet("color:#6B778C; font-size:11px; font-weight:600; background:transparent;")
            return l

        def _lbl_val(text):
            l = QLabel(text)
            l.setStyleSheet("color:#172B4D; font-size:13px; background:transparent;")
            return l

        hf_layout.addWidget(_lbl_key("ДОКУМЕНТ"),       0, 0)
        hf_layout.addWidget(_lbl_val(op_data["doc_name"]), 1, 0)
        hf_layout.addWidget(_lbl_key("ОПЕРАЦИЯ"),       0, 1)
        hf_layout.addWidget(_lbl_val(op_type),          1, 1)
        hf_layout.addWidget(_lbl_key("ДАТА"),           0, 2)
        hf_layout.addWidget(_lbl_val(op_data["date"]),  1, 2)
        hf_layout.addWidget(_lbl_key("ПОДРАЗДЕЛЕНИЕ"),  0, 3)
        hf_layout.addWidget(_lbl_val(op_data["unit_name"] or "—"), 1, 3)
        layout.addWidget(header_frame)

        # ── Таблица позиций ───────────────────────────────────────────────────
        pos_label = QLabel(f"Позиции ({len(op_data['rows'])}):")
        pos_label.setStyleSheet("font-weight:600; font-size:13px; color:#172B4D;")
        layout.addWidget(pos_label)

        table = QTableWidget(0, 4)
        table.setHorizontalHeaderLabels(["Название", "Размер", "Н/Н (полный)", "Кол-во / S/N"])
        hh = table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.verticalHeader().setVisible(False)

        for r in op_data["rows"]:
            row_idx = table.rowCount()
            table.insertRow(row_idx)
            qty_sn = r["factory_sn"] if r["factory_sn"] else str(r["quantity"] or "")
            table.setItem(row_idx, 0, QTableWidgetItem(r["item_name"]))
            table.setItem(row_idx, 1, QTableWidgetItem(r["size_name"] or ""))
            table.setItem(row_idx, 2, QTableWidgetItem(r["full_code"] or ""))
            table.setItem(row_idx, 3, QTableWidgetItem(qty_sn))

        layout.addWidget(table, 1)

        # ── Кнопка закрытия ──────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        close_btn = QPushButton("Закрыть")
        close_btn.setFixedWidth(120)
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)


class JournalTab(QWidget):
    """Вкладка журнала операций: плоский список, двойной клик открывает документ."""

    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self._ops: list[dict] = []   # сгруппированные операции для открытия диалога
        self._build_ui()
        self.load_units()
        self.load_journal()

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(24, 20, 24, 20)
        main_layout.setSpacing(12)

        hint = QLabel("Двойной клик по строке — открыть документ операции")
        hint.setStyleSheet("color:#6B778C; font-size:12px;")
        main_layout.addWidget(hint)

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Дата с:"))
        self.date_from_edit = QDateEdit()
        self.date_from_edit.setCalendarPopup(True)
        self.date_from_edit.setDisplayFormat("dd.MM.yyyy")
        self.date_from_edit.setDate(QDate.currentDate().addMonths(-1))
        self.date_from_edit.setFixedWidth(120)
        _patch_calendar_arrows(self.date_from_edit)
        filter_layout.addWidget(self.date_from_edit)
        filter_layout.addWidget(QLabel("по:"))
        self.date_to_edit = QDateEdit()
        self.date_to_edit.setCalendarPopup(True)
        self.date_to_edit.setDisplayFormat("dd.MM.yyyy")
        self.date_to_edit.setDate(QDate.currentDate())
        self.date_to_edit.setFixedWidth(120)
        _patch_calendar_arrows(self.date_to_edit)
        filter_layout.addWidget(self.date_to_edit)
        filter_layout.addWidget(QLabel("Подразделение:"))
        self.filter_unit_combo = QComboBox()
        self.filter_unit_combo.setEditable(False)
        filter_layout.addWidget(self.filter_unit_combo)
        self.filter_btn = QPushButton("Применить")
        filter_layout.addWidget(self.filter_btn)
        filter_layout.addStretch()
        self.export_journal_excel_btn = QPushButton()
        self.export_journal_excel_btn.setObjectName("ExportBtn")
        self.export_journal_excel_btn.setIcon(_icon_from_file("microsoft-excel-logo-duotone.svg", 18))
        self.export_journal_excel_btn.setIconSize(QSize(18, 18))
        self.export_journal_excel_btn.setFixedSize(32, 32)
        self.export_journal_excel_btn.setToolTip("Экспорт в Excel")
        self.export_journal_pdf_btn = QPushButton()
        self.export_journal_pdf_btn.setObjectName("ExportBtn")
        self.export_journal_pdf_btn.setIcon(_icon_from_file("file-pdf-duotone.svg", 18))
        self.export_journal_pdf_btn.setIconSize(QSize(18, 18))
        self.export_journal_pdf_btn.setFixedSize(32, 32)
        self.export_journal_pdf_btn.setToolTip("Экспорт в PDF")
        filter_layout.addWidget(self.export_journal_excel_btn)
        filter_layout.addWidget(self.export_journal_pdf_btn)
        main_layout.addLayout(filter_layout)

        # Плоская таблица — одна строка на операцию
        self.journal_table = QTableWidget(0, 5)
        self.journal_table.setHorizontalHeaderLabels(
            ["Дата", "Операция", "Документ", "Позиций", "Подразделение"]
        )
        hh = self.journal_table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.journal_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.journal_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.journal_table.verticalHeader().setVisible(False)
        self.journal_table.cellDoubleClicked.connect(self._on_row_double_click)
        main_layout.addWidget(self.journal_table, 1)

        self.filter_btn.clicked.connect(self.load_journal)
        self.export_journal_excel_btn.clicked.connect(self.on_export_journal_excel)
        self.export_journal_pdf_btn.clicked.connect(self.on_export_journal_pdf)

    def load_units(self):
        units = self.db.get_units()
        self.filter_unit_combo.clear()
        self.filter_unit_combo.addItem("— Все —", None)
        for u in units:
            self.filter_unit_combo.addItem(u["name"], u["id"])

    def _on_row_double_click(self, row: int, _col: int):
        if 0 <= row < len(self._ops):
            dlg = OperationDetailDialog(self._ops[row], self)
            dlg.exec()

    def load_journal(self):
        from collections import OrderedDict
        date_from = self.date_from_edit.date().toString("yyyy-MM-dd")
        date_to   = self.date_to_edit.date().toString("yyyy-MM-dd")
        unit_id   = self.filter_unit_combo.currentData()
        rows = self.db.get_journal_view(date_from=date_from, date_to=date_to, unit_id=unit_id)

        # Группируем по (doc_name, дата-без-времени, op_type, unit_name)
        groups: OrderedDict = OrderedDict()
        for row in rows:
            date_only = (row["date"] or "")[:10]
            key = (row["doc_name"] or "", date_only, row["op_type"], row["unit_name"] or "")
            if key not in groups:
                groups[key] = {
                    "doc_name":  row["doc_name"] or "",
                    "date":      date_only,
                    "op_type":   row["op_type"],
                    "unit_name": row["unit_name"] or "",
                    "rows":      [],
                }
            groups[key]["rows"].append(row)

        self._ops = list(groups.values())
        self.journal_table.setRowCount(0)

        for i, op in enumerate(self._ops):
            op_text = "Приход" if op["op_type"] == "IN" else "Выдача"
            self.journal_table.insertRow(i)
            self.journal_table.setItem(i, 0, QTableWidgetItem(op["date"]))
            self.journal_table.setItem(i, 1, QTableWidgetItem(op_text))
            self.journal_table.setItem(i, 2, QTableWidgetItem(op["doc_name"]))
            self.journal_table.setItem(i, 3, QTableWidgetItem(str(len(op["rows"]))))
            self.journal_table.setItem(i, 4, QTableWidgetItem(op["unit_name"]))

    def on_export_journal_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Экспорт журнала", "", "Excel (*.xlsx)")
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"
        date_from = self.date_from_edit.date().toString("yyyy-MM-dd")
        date_to = self.date_to_edit.date().toString("yyyy-MM-dd")
        unit_id = self.filter_unit_combo.currentData()
        if _export_journal_excel(self.db, path, date_from, date_to, unit_id):
            QMessageBox.information(self, "Экспорт", "Журнал экспортирован в Excel.")
        else:
            QMessageBox.warning(self, "Ошибка", "Не удалось экспортировать. Установите: pip install openpyxl")

    def on_export_journal_pdf(self):
        path, _ = QFileDialog.getSaveFileName(self, "Экспорт журнала", "", "PDF (*.pdf)")
        if not path:
            return
        if not path.endswith(".pdf"):
            path += ".pdf"
        date_from = self.date_from_edit.date().toString("yyyy-MM-dd")
        date_to = self.date_to_edit.date().toString("yyyy-MM-dd")
        unit_id = self.filter_unit_combo.currentData()
        if _export_journal_pdf(self.db, path, date_from, date_to, unit_id):
            QMessageBox.information(self, "Экспорт", "Журнал экспортирован в PDF.")
        else:
            QMessageBox.warning(self, "Ошибка", "Не удалось экспортировать. Установите: pip install reportlab")


class BasketDialog(QDialog):
    """Модальное окно корзины: просмотр позиций, редактирование, проведение операции."""

    def __init__(
        self,
        basket: list[dict],
        db: DatabaseManager,
        op_type: str,
        units: list,
        parent=None,
    ):
        super().__init__(parent)
        self._basket = basket
        self.db = db
        self._op_type = op_type
        self._units = units

        self.setWindowTitle("Корзина операции")
        self.setMinimumWidth(780)
        self.setMinimumHeight(560)
        self._build_ui()
        self._refresh()

    def _build_ui(self):
        if self._op_type == "IN":
            op_text        = "ПРИХОД"
            badge_bg       = "#E3FCEF"
            badge_fg       = "#006644"
            badge_border   = "#ABF5D1"
            post_btn_color = "#00875A"
            post_btn_hover = "#006644"
        else:
            op_text        = "ВЫДАЧА"
            badge_bg       = "#FFF4E6"
            badge_fg       = "#974F0C"
            badge_border   = "#FFE380"
            post_btn_color = "#FF8B00"
            post_btn_hover = "#974F0C"

        self.setStyleSheet(
            f"""
            QDialog {{
                background: #FFFFFF;
            }}
            QLabel {{
                background: transparent;
                color: #172B4D;
            }}
            QLineEdit, QComboBox {{
                background: #FFFFFF;
                border: 2px solid #DFE1E6;
                border-radius: 3px;
                padding: 0 8px;
                color: #172B4D;
                min-height: 32px;
                max-height: 32px;
            }}
            QLineEdit:focus, QComboBox:focus {{
                border-color: #4C9AFF;
            }}
            QLineEdit:hover, QComboBox:hover {{
                border-color: #B3BAC5;
            }}
            QComboBox::drop-down {{ border: none; padding-right: 8px; }}
            QComboBox QAbstractItemView {{
                background: #FFFFFF;
                border: 1px solid #DFE1E6;
                selection-background-color: #DEEBFF;
                color: #172B4D;
            }}
            QTableWidget {{
                background: #FFFFFF;
                border: 1px solid #DFE1E6;
                border-radius: 4px;
                gridline-color: #F4F5F7;
                outline: none;
                color: #172B4D;
            }}
            QTableWidget::item {{
                padding: 0 12px;
                border-bottom: 1px solid #F4F5F7;
                color: #172B4D;
            }}
            QTableWidget::item:selected {{
                background: #DEEBFF;
                color: #172B4D;
            }}
            QHeaderView::section {{
                background: #F4F5F7;
                border: none;
                border-bottom: 2px solid #DFE1E6;
                border-right: 1px solid #DFE1E6;
                padding: 8px 12px;
                font-weight: 600;
                font-size: 11px;
                color: #6B778C;
            }}
            QPushButton#PostBtn {{
                background: {post_btn_color};
                color: #FFFFFF;
                border: none;
                border-radius: 3px;
                padding: 0 20px;
                font-weight: 600;
                min-height: 36px;
                max-height: 36px;
                font-size: 13px;
            }}
            QPushButton#PostBtn:hover {{ background: {post_btn_hover}; }}
            QPushButton#PostBtn:disabled {{
                background: #DFE1E6; color: #97A0AF;
            }}
            QPushButton#CloseBtn {{
                background: transparent;
                color: #42526E;
                border: 1.5px solid #DFE1E6;
                border-radius: 3px;
                padding: 0 16px;
                min-height: 36px;
                max-height: 36px;
                font-size: 13px;
            }}
            QPushButton#CloseBtn:hover {{
                background: #F4F5F7;
                border-color: #B3BAC5;
            }}
            QPushButton#ClearBtn {{
                background: transparent;
                color: #DE350B;
                border: 1.5px solid #DE350B;
                border-radius: 3px;
                padding: 0 16px;
                min-height: 36px;
                max-height: 36px;
                font-size: 13px;
            }}
            QPushButton#ClearBtn:hover {{ background: #FFEBE6; }}
            QPushButton#ClearBtn:disabled {{ color: #DFE1E6; border-color: #DFE1E6; }}
            """
        )

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ──────────────────────────────────────────────────────────────
        # 1. Цветная шапка
        # ──────────────────────────────────────────────────────────────
        header_frame = QFrame()
        header_frame.setFixedHeight(72)
        header_frame.setStyleSheet(
            f"QFrame {{ background: {badge_bg};"
            f" border-bottom: 1px solid {badge_border}; }}"
        )
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(24, 0, 24, 0)
        header_layout.setSpacing(14)

        badge = QLabel(op_text)
        badge.setFixedHeight(26)
        badge.setStyleSheet(
            f"color: {badge_fg}; background: {badge_border};"
            f" border-radius: 4px; padding: 0 10px;"
            f" font-weight: 700; font-size: 11px;"
        )
        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)

        title_lbl = QLabel("Корзина операции")
        title_lbl.setStyleSheet(
            "font-size: 18px; font-weight: 700; color: #172B4D;"
        )

        self.count_lbl = QLabel("")
        self.count_lbl.setStyleSheet(
            "font-size: 13px; color: #6B778C;"
        )

        header_layout.addWidget(badge)
        header_layout.addWidget(title_lbl)
        header_layout.addWidget(self.count_lbl)
        header_layout.addStretch()
        root.addWidget(header_frame)

        # ──────────────────────────────────────────────────────────────
        # 2. Основной контент (таблица / empty state)
        # ──────────────────────────────────────────────────────────────
        content = QWidget()
        content.setStyleSheet("QWidget { background: #FFFFFF; }")
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(24, 20, 24, 16)
        content_layout.setSpacing(16)

        # Таблица
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Название", "Размер", "Н/Н (полный)", "Кол-во / S/N", ""])
        h = self.table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(4, 40)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        content_layout.addWidget(self.table, 1)

        # Empty state
        self.empty_widget = QFrame()
        self.empty_widget.setStyleSheet(
            "QFrame { background: #F4F5F7; border: 1px dashed #B3BAC5; border-radius: 6px; }"
        )
        empty_layout = QVBoxLayout(self.empty_widget)
        empty_layout.setContentsMargins(0, 40, 0, 40)
        empty_lbl = QLabel("Корзина пуста")
        empty_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_lbl.setStyleSheet("font-size: 15px; color: #97A0AF; font-weight: 600; border: none;")
        sub_lbl = QLabel("Добавьте товары из поиска")
        sub_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub_lbl.setStyleSheet("font-size: 12px; color: #B3BAC5; border: none;")
        empty_layout.addWidget(empty_lbl)
        empty_layout.addWidget(sub_lbl)
        content_layout.addWidget(self.empty_widget)

        # Форма: документ + подразделение
        form_frame = QFrame()
        form_frame.setStyleSheet(
            "QFrame { background: #F4F5F7; border: 1px solid #DFE1E6; border-radius: 4px; }"
        )
        form_frame.setFixedHeight(78)
        form_layout = QHBoxLayout(form_frame)
        form_layout.setContentsMargins(16, 10, 16, 10)
        form_layout.setSpacing(24)
        form_layout.setAlignment(Qt.AlignmentFlag.AlignVCenter)

        doc_col = QVBoxLayout()
        doc_col.setSpacing(4)
        doc_col.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        doc_lbl = QLabel("Документ *")
        doc_lbl.setStyleSheet("font-size: 11px; font-weight: 700; color: #6B778C; border: none;")
        self.doc_edit = QLineEdit()
        self.doc_edit.setPlaceholderText("Введите № учётного документа")
        doc_col.addWidget(doc_lbl)
        doc_col.addWidget(self.doc_edit)

        unit_col = QVBoxLayout()
        unit_col.setSpacing(4)
        unit_col.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        unit_lbl_text = "От подразделения *" if self._op_type == "IN" else "В подразделение *"
        unit_lbl = QLabel(unit_lbl_text)
        unit_lbl.setStyleSheet("font-size: 11px; font-weight: 700; color: #6B778C; border: none;")
        self.unit_combo = QComboBox()
        self.unit_combo.setEditable(False)
        self.unit_combo.setMinimumWidth(200)
        for u in self._units:
            self.unit_combo.addItem(u["name"], u["id"])
        unit_col.addWidget(unit_lbl)
        unit_col.addWidget(self.unit_combo)

        form_layout.addLayout(doc_col, 1)
        form_layout.addLayout(unit_col)
        content_layout.addWidget(form_frame)

        root.addWidget(content, 1)

        # ──────────────────────────────────────────────────────────────
        # 3. Нижняя панель с кнопками
        # ──────────────────────────────────────────────────────────────
        footer_frame = QFrame()
        footer_frame.setFixedHeight(64)
        footer_frame.setStyleSheet(
            "QFrame { background: #F4F5F7; border-top: 1px solid #DFE1E6; }"
        )
        footer_layout = QHBoxLayout(footer_frame)
        footer_layout.setContentsMargins(24, 0, 24, 0)
        footer_layout.setSpacing(10)

        self.clear_btn = QPushButton("Очистить")
        self.clear_btn.setObjectName("ClearBtn")
        self.post_btn = QPushButton("Провести операцию →")
        self.post_btn.setObjectName("PostBtn")
        close_btn = QPushButton("Отмена")
        close_btn.setObjectName("CloseBtn")

        footer_layout.addWidget(self.clear_btn)
        footer_layout.addStretch()
        footer_layout.addWidget(close_btn)
        footer_layout.addWidget(self.post_btn)
        root.addWidget(footer_frame)

        self.clear_btn.clicked.connect(self._on_clear)
        self.post_btn.clicked.connect(self._on_post)
        close_btn.clicked.connect(self.reject)

    def _refresh(self):
        self.table.setRowCount(0)
        has_items = bool(self._basket)
        self.table.setVisible(has_items)
        self.empty_widget.setVisible(not has_items)
        self.post_btn.setEnabled(has_items)
        self.clear_btn.setEnabled(has_items)

        n = len(self._basket)
        self.count_lbl.setText(f"— {n} поз." if n else "")

        for i, pos in enumerate(self._basket):
            self.table.insertRow(i)
            self.table.setRowHeight(i, 40)
            self.table.setItem(i, 0, QTableWidgetItem(pos["item_name"]))
            self.table.setItem(i, 1, QTableWidgetItem(pos["size_name"] or "UNI"))
            self.table.setItem(i, 2, QTableWidgetItem(pos["full_code"] or ""))
            val = pos["sn"] if pos["item_type"] == "serial" else str(pos["qty"])
            qty_item = QTableWidgetItem(val)
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(i, 3, qty_item)
            del_btn = QPushButton("✕")
            del_btn.setFixedSize(28, 28)
            del_btn.setStyleSheet(
                "QPushButton { background: transparent; color: #97A0AF; border: 1px solid #DFE1E6;"
                " border-radius: 14px; font-size: 12px; padding: 0; }"
                "QPushButton:hover { background: #FF5630; color: #FFFFFF; border-color: #FF5630; }"
            )
            del_btn.clicked.connect(lambda _, idx=i: self._remove_item(idx))
            container = QWidget()
            container.setStyleSheet("QWidget { background: transparent; }")
            c_layout = QHBoxLayout(container)
            c_layout.setContentsMargins(0, 0, 0, 0)
            c_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            c_layout.addWidget(del_btn)
            self.table.setCellWidget(i, 4, container)

    def _remove_item(self, idx: int):
        if 0 <= idx < len(self._basket):
            self._basket.pop(idx)
            self._refresh()

    def _on_clear(self):
        if not self._basket:
            return
        reply = QMessageBox.question(
            self, "Очистить корзину", "Удалить все позиции из корзины?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._basket.clear()
            self._refresh()

    def _on_post(self):
        if not self._basket:
            QMessageBox.warning(self, "Ошибка", "Корзина пуста.")
            return

        doc_name = self.doc_edit.text().strip()
        if not doc_name:
            QMessageBox.warning(self, "Ошибка", "Введите номер учётного документа.")
            self.doc_edit.setFocus()
            return

        unit_id = self.unit_combo.currentData()
        if unit_id is None:
            QMessageBox.warning(self, "Ошибка", "Выберите подразделение.")
            return

        # ── Предварительная проверка ──
        errors = []
        for pos in self._basket:
            vid = pos["variant_id"]
            if pos["item_type"] == "qty":
                if self._op_type == "OUT":
                    current = self.db.get_qty_stock(vid)
                    if current < pos["qty"]:
                        errors.append(
                            f"«{pos['item_name']} / {pos['size_name']}»: "
                            f"запрошено {pos['qty']}, на складе {current}"
                        )
            else:
                sn = pos["sn"]
                if self._op_type == "IN":
                    if self.db.serial_exists(sn):
                        errors.append(f"S/N «{sn}» уже числится на складе")
                else:
                    if not self.db.serial_exists_for_variant(vid, sn):
                        errors.append(f"S/N «{sn}» не найден на складе")

        if errors:
            QMessageBox.warning(
                self, "Ошибки в позициях",
                "Невозможно провести операцию:\n\n" + "\n".join(f"• {e}" for e in errors),
            )
            return

        # ── Проводим ──
        try:
            for pos in self._basket:
                vid = pos["variant_id"]
                if pos["item_type"] == "qty":
                    delta = pos["qty"] if self._op_type == "IN" else -pos["qty"]
                    self.db.adjust_qty_stock(vid, delta)
                    self.db.add_journal_record(self._op_type, vid, pos["qty"], None, unit_id, doc_name)
                else:
                    sn = pos["sn"]
                    if self._op_type == "IN":
                        self.db.add_serial(vid, sn)
                    else:
                        self.db.remove_serial(vid, sn)
                    self.db.add_journal_record(self._op_type, vid, 1, sn, unit_id, doc_name)
        except sqlite3.IntegrityError as e:
            if logger:
                logger.exception("Post operation failed: %s", e)
            QMessageBox.warning(self, "Ошибка БД", f"Не удалось провести операцию:\n{e}")
            return

        n = len(self._basket)
        QMessageBox.information(self, "Успех", f"Операция проведена: {n} поз.")
        self.accept()


class OperationsTab(QWidget):
    """Вкладка проведения операций: поддерживает несколько наименований в одной операции."""

    # Структура позиции в корзине
    # {variant_id, item_name, full_code, size_name, item_type, qty, sn}

    def __init__(self, db: DatabaseManager, stock_tab_updater, refresh_journal=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.stock_tab_updater = stock_tab_updater
        self.refresh_journal = refresh_journal
        self.selected_variant = None
        self._search_rows = []
        self._basket: list[dict] = []
        self._units: list = []
        self._build_ui()
        self.load_units()

    # ── UI ──────────────────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(12)

        # ── Строка: тип операции + кнопка Корзина ──
        op_bar = QHBoxLayout()
        self.in_radio = QRadioButton("ПРИХОД")
        self.out_radio = QRadioButton("ВЫДАЧА")
        self.in_radio.setChecked(True)
        self.op_group = QButtonGroup(self)
        self.op_group.addButton(self.in_radio)
        self.op_group.addButton(self.out_radio)
        op_bar.addWidget(self.in_radio)
        op_bar.addWidget(self.out_radio)
        op_bar.addStretch()
        self.basket_btn = QPushButton("Корзина (0)")
        self.basket_btn.setObjectName("BasketBtn")
        self.basket_btn.setMinimumWidth(130)
        op_bar.addWidget(self.basket_btn)
        root.addLayout(op_bar)

        # ── Поиск ──
        search_row = QHBoxLayout()
        search_row.addWidget(QLabel("Поиск:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Название или Н/Н код...")
        self.search_btn = QPushButton("Найти")
        search_row.addWidget(self.search_edit, 1)
        search_row.addWidget(self.search_btn)
        root.addLayout(search_row)

        self.results_table = QTableWidget(0, 5)
        self.results_table.setHorizontalHeaderLabels(["Код", "Название", "Размер", "Категория", "Наличие"])
        hh = self.results_table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.results_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.results_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        root.addWidget(self.results_table, 1)

        # ── Панель добавления выбранного товара ──
        add_panel = QFrame()
        add_panel.setObjectName("AddPanel")
        add_layout = QVBoxLayout(add_panel)
        add_layout.setContentsMargins(12, 10, 12, 10)
        add_layout.setSpacing(8)

        self.selected_label = QLabel("Выберите товар из результатов поиска")
        self.selected_label.setObjectName("SelectedLabel")
        self.selected_label.setWordWrap(True)
        add_layout.addWidget(self.selected_label)

        qty_sn_row = QHBoxLayout()
        qty_sn_row.setSpacing(8)

        # Лейбл + спинбокс в плотном контейнере
        self.qty_label = QLabel("Количество:")
        self.qty_spin = QSpinBox()
        self.qty_spin.setRange(1, 1_000_000)
        self.qty_spin.setValue(1)
        self.qty_spin.setFixedWidth(80)
        qty_pair = QWidget()
        qty_pair.setStyleSheet("QWidget { background: transparent; }")
        qty_pair_layout = QHBoxLayout(qty_pair)
        qty_pair_layout.setContentsMargins(0, 0, 0, 0)
        qty_pair_layout.setSpacing(6)
        qty_pair_layout.addWidget(self.qty_label)
        qty_pair_layout.addWidget(self.qty_spin)

        self.sn_label = QLabel("S/N:")
        self.sn_edit = QLineEdit()
        self.sn_edit.setPlaceholderText("Заводской номер")
        self.sn_combo = QComboBox()
        self.sn_combo.setEditable(False)
        self.sn_combo.setMinimumWidth(200)
        self.add_btn = QPushButton("+ Добавить в корзину")
        self.add_btn.setEnabled(False)
        self.add_btn.setMinimumWidth(180)

        qty_sn_row.addWidget(qty_pair)
        qty_sn_row.addSpacing(4)
        qty_sn_row.addWidget(self.sn_label)
        qty_sn_row.addWidget(self.sn_edit, 1)
        qty_sn_row.addWidget(self.sn_combo, 1)
        qty_sn_row.addSpacing(8)
        qty_sn_row.addWidget(self.add_btn)
        add_layout.addLayout(qty_sn_row)
        root.addWidget(add_panel)

        # Стили
        add_panel.setStyleSheet(
            "QFrame#AddPanel { background:#FFFFFF; border:1px solid #DFE1E6; border-radius:3px; }"
        )
        self.selected_label.setStyleSheet("color:#6B778C; font-size:12px; background:transparent;")

        # Сигналы
        self.search_btn.clicked.connect(self.on_search)
        self.search_edit.returnPressed.connect(self.on_search)
        self.search_edit.textChanged.connect(self._on_search_text_changed)
        self.results_table.cellClicked.connect(self.on_result_clicked)
        self.add_btn.clicked.connect(self.on_add_to_basket)
        self.basket_btn.clicked.connect(self._open_basket)
        self.in_radio.toggled.connect(self._on_op_type_changed)
        self.out_radio.toggled.connect(self._on_op_type_changed)

        self._set_input_mode(None)
        self._update_basket_btn()

    # ── Helpers ─────────────────────────────────────────────────────────────

    def load_units(self):
        self._units = list(self.db.get_units())

    def _set_input_mode(self, item_type: str | None):
        is_qty   = item_type == "qty"
        is_sn    = item_type == "serial"
        is_out   = self.out_radio.isChecked()

        # Количество — только для Мат. средств
        self.qty_label.setVisible(is_qty or item_type is None)
        self.qty_spin.setVisible(is_qty or item_type is None)
        self.qty_label.setEnabled(is_qty)
        self.qty_spin.setEnabled(is_qty)

        # S/N: при ВЫДАЧЕ — комбо, при ПРИХОДЕ — текстовый ввод
        use_combo = is_sn and is_out
        self.sn_label.setVisible(is_sn or item_type is None)
        self.sn_label.setEnabled(is_sn)
        self.sn_edit.setVisible(is_sn and not use_combo)
        self.sn_edit.setEnabled(is_sn and not use_combo)
        self.sn_combo.setVisible(use_combo)
        self.sn_combo.setEnabled(use_combo)

        self.add_btn.setEnabled(item_type is not None)

    def _update_basket_btn(self):
        n = len(self._basket)
        self.basket_btn.setText(f"Корзина ({n})")
        if n > 0:
            self.basket_btn.setStyleSheet(
                "QPushButton#BasketBtn { background:#0052CC; color:#FFFFFF;"
                " border:none; border-radius:3px; font-weight:600; }"
                "QPushButton#BasketBtn:hover { background:#0747A6; }"
            )
        else:
            self.basket_btn.setStyleSheet(
                "QPushButton#BasketBtn { background:transparent; color:#6B778C;"
                " border:1px solid #B3BAC5; border-radius:3px; font-weight:400; }"
                "QPushButton#BasketBtn:hover { background:#F4F5F7; }"
            )

    def _open_basket(self):
        op_type = "IN" if self.in_radio.isChecked() else "OUT"
        dlg = BasketDialog(
            basket=self._basket,
            db=self.db,
            op_type=op_type,
            units=self._units,
            parent=self,
        )
        result = dlg.exec()
        self._update_basket_btn()

        if result == QDialog.DialogCode.Accepted:
            self._basket.clear()
            self._update_basket_btn()
            self.search_edit.clear()
            self.results_table.setRowCount(0)
            self._search_rows = []
            self.selected_variant = None
            self.selected_label.setText("Выберите товар из результатов поиска")
            self._set_input_mode(None)
            self.sn_edit.clear()
            self.qty_spin.setValue(1)
            self.load_units()
            if callable(self.refresh_journal):
                self.refresh_journal()
            self.stock_tab_updater()

    # ── Slots ────────────────────────────────────────────────────────────────

    def _do_search(self, text: str = ""):
        """Основная логика поиска. Вызывается явно или при автофильтрации."""
        is_out = self.out_radio.isChecked()
        rows = self.db.search_variants(text, only_in_stock=is_out)
        self.results_table.setRowCount(0)
        self._search_rows = rows

        for i, row in enumerate(rows):
            self.results_table.insertRow(i)
            self.results_table.setItem(i, 0, QTableWidgetItem(row["full_code"] or ""))
            self.results_table.setItem(i, 1, QTableWidgetItem(row["item_name"] or ""))
            self.results_table.setItem(i, 2, QTableWidgetItem(row["size_name"] or ""))
            self.results_table.setItem(i, 3, QTableWidgetItem(row["category_name"] or "—"))

            stock_val = row["stock_value"] or 0
            stock_item = QTableWidgetItem(str(stock_val))
            stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if is_out:
                stock_item.setForeground(QColor("#006644") if stock_val > 0 else QColor("#DE350B"))
            else:
                stock_item.setForeground(QColor("#97A0AF"))
            self.results_table.setItem(i, 4, stock_item)

        self.selected_variant = None
        self.selected_label.setText("Выберите товар из результатов поиска")
        self._set_input_mode(None)

    def on_search(self):
        """Поиск по кнопке / Enter — работает в обоих режимах."""
        text = self.search_edit.text().strip()
        if not text and self.in_radio.isChecked():
            QMessageBox.warning(self, "Ошибка", "Введите текст для поиска.")
            return
        self._do_search(text)

    def _on_search_text_changed(self, text: str):
        """Живая фильтрация при вводе — только в режиме ВЫДАЧА."""
        if self.out_radio.isChecked():
            self._do_search(text.strip())

    def _on_op_type_changed(self):
        """При переключении типа операции — обновляем таблицу."""
        if self.out_radio.isChecked():
            # ВЫДАЧА: сразу показываем весь склад
            self._do_search(self.search_edit.text().strip())
        else:
            # ПРИХОД: очищаем, ждём ввода от пользователя
            self.results_table.setRowCount(0)
            self._search_rows = []
            self.selected_variant = None
            self.selected_label.setText("Выберите товар из результатов поиска")
            self._set_input_mode(None)

    def on_result_clicked(self, row: int, _col: int):
        if row < 0 or row >= len(self._search_rows):
            return
        vrow = self._search_rows[row]
        self.selected_variant = vrow
        type_text = "Мат. средства" if vrow["item_type"] == "qty" else "Основные средства"
        self.selected_label.setText(
            f"{vrow['item_name']}  |  {vrow['size_name']}  |  {vrow['full_code']}  ({type_text})"
        )
        self._set_input_mode(vrow["item_type"])
        if vrow["item_type"] == "qty":
            self.qty_spin.setValue(1)
        elif self.out_radio.isChecked():
            # ВЫДАЧА + серийный: заполняем комбо доступными S/N
            serials = self.db.get_serials_for_variant(vrow["variant_id"])
            self.sn_combo.clear()
            for s in serials:
                self.sn_combo.addItem(s["factory_sn"])
            if self.sn_combo.count() == 0:
                self.sn_combo.addItem("— нет на складе —")
                self.add_btn.setEnabled(False)
        else:
            self.sn_edit.clear()

    def on_add_to_basket(self):
        if self.selected_variant is None:
            return
        vrow = self.selected_variant
        item_type = vrow["item_type"]

        if item_type == "qty":
            qty = self.qty_spin.value()
            if qty <= 0:
                QMessageBox.warning(self, "Ошибка", "Количество должно быть больше 0.")
                return
            self._basket.append({
                "variant_id": vrow["variant_id"],
                "item_name":  vrow["item_name"],
                "full_code":  vrow["full_code"],
                "size_name":  vrow["size_name"],
                "item_type":  "qty",
                "qty":        qty,
                "sn":         None,
            })
            self.qty_spin.setValue(1)
        else:
            # Серийный: источник S/N зависит от режима
            if self.out_radio.isChecked():
                sn = self.sn_combo.currentText().strip()
            else:
                sn = self.sn_edit.text().strip()

            if not sn or sn == "— нет на складе —":
                QMessageBox.warning(self, "Ошибка", "Выберите или введите заводской номер (S/N).")
                return
            if any(p["sn"] == sn for p in self._basket):
                QMessageBox.warning(self, "Ошибка", f"S/N «{sn}» уже добавлен в эту операцию.")
                return
            self._basket.append({
                "variant_id": vrow["variant_id"],
                "item_name":  vrow["item_name"],
                "full_code":  vrow["full_code"],
                "size_name":  vrow["size_name"],
                "item_type":  "serial",
                "qty":        1,
                "sn":         sn,
            })
            if self.out_radio.isChecked():
                # Убираем использованный S/N из комбо
                self.sn_combo.removeItem(self.sn_combo.currentIndex())
                if self.sn_combo.count() == 0:
                    self.add_btn.setEnabled(False)
            else:
                self.sn_edit.clear()

        self._update_basket_btn()


class StockTab(QWidget):
    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self._group_by_category = False
        self._build_ui()
        self.reload()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(12)

        # Поиск по складу и экспорт
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Поиск:"))
        self.search_edit = QLineEdit()
        self.search_btn = QPushButton("Найти")
        search_layout.addWidget(self.search_edit)
        search_layout.addWidget(self.search_btn)

        # Переключатель вид
        self.view_all_btn = QPushButton("Все остатки")
        self.view_cat_btn = QPushButton("По категориям")
        self.view_all_btn.setCheckable(True)
        self.view_cat_btn.setCheckable(True)
        self.view_all_btn.setChecked(True)
        self.view_all_btn.clicked.connect(self._on_view_all)
        self.view_cat_btn.clicked.connect(self._on_view_cat)
        search_layout.addWidget(self.view_all_btn)
        search_layout.addWidget(self.view_cat_btn)

        search_layout.addStretch()
        self.export_stock_excel_btn = QPushButton()
        self.export_stock_excel_btn.setObjectName("ExportBtn")
        self.export_stock_excel_btn.setIcon(_icon_from_file("microsoft-excel-logo-duotone.svg", 18))
        self.export_stock_excel_btn.setIconSize(QSize(18, 18))
        self.export_stock_excel_btn.setFixedSize(32, 32)
        self.export_stock_excel_btn.setToolTip("Экспорт в Excel")
        self.export_stock_pdf_btn = QPushButton()
        self.export_stock_pdf_btn.setObjectName("ExportBtn")
        self.export_stock_pdf_btn.setIcon(_icon_from_file("file-pdf-duotone.svg", 18))
        self.export_stock_pdf_btn.setIconSize(QSize(18, 18))
        self.export_stock_pdf_btn.setFixedSize(32, 32)
        self.export_stock_pdf_btn.setToolTip("Экспорт в PDF")
        search_layout.addWidget(self.export_stock_excel_btn)
        search_layout.addWidget(self.export_stock_pdf_btn)
        layout.addLayout(search_layout)

        self.tree = QTreeWidget()
        self.tree.setColumnCount(5)
        self.tree.setHeaderLabels(
            ["Н/Н (базовый)", "Название", "Размер", "Остаток", "Ед. изм."]
        )
        header = self.tree.header()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        header.setStretchLastSection(False)
        self.tree.setColumnWidth(0, 150)
        self.tree.setColumnWidth(2, 100)
        self.tree.setColumnWidth(3, 80)
        self.tree.setColumnWidth(4, 70)
        self.tree.setRootIsDecorated(False)
        self.tree.setIndentation(16)
        self.tree.setUniformRowHeights(True)
        self.tree.setExpandsOnDoubleClick(True)
        self.tree.setAlternatingRowColors(False)
        self.tree.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tree.itemClicked.connect(self.on_item_clicked)

        self.search_btn.clicked.connect(self.on_search)
        self.search_edit.returnPressed.connect(self.on_search)
        self.export_stock_excel_btn.clicked.connect(self.on_export_stock_excel)
        self.export_stock_pdf_btn.clicked.connect(self.on_export_stock_pdf)

        layout.addWidget(self.tree)

    def _on_view_all(self):
        self._group_by_category = False
        self.view_all_btn.setChecked(True)
        self.view_cat_btn.setChecked(False)
        self.reload()

    def _on_view_cat(self):
        self._group_by_category = True
        self.view_cat_btn.setChecked(True)
        self.view_all_btn.setChecked(False)
        self.reload()

    def on_export_stock_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Экспорт остатков", "", "Excel (*.xlsx)")
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"
        if _export_stock_excel(self.db, path):
            QMessageBox.information(self, "Экспорт", "Остатки экспортированы в Excel.")
        else:
            QMessageBox.warning(self, "Ошибка", "Не удалось экспортировать. Установите: pip install openpyxl")

    def on_export_stock_pdf(self):
        path, _ = QFileDialog.getSaveFileName(self, "Экспорт остатков", "", "PDF (*.pdf)")
        if not path:
            return
        if not path.endswith(".pdf"):
            path += ".pdf"
        if _export_stock_pdf(self.db, path):
            QMessageBox.information(self, "Экспорт", "Остатки экспортированы в PDF.")
        else:
            QMessageBox.warning(self, "Ошибка", "Не удалось экспортировать. Установите: pip install reportlab")

    def _make_item_node(self, parent, row):
        """Добавить узел изделия, дочерние размеры и (для серийных) S/N-записи."""
        details = self.db.get_item_stock_details(row["item_id"])
        filter_text = self.search_edit.text().strip().lower() if hasattr(self, "search_edit") else ""
        is_serial = row["item_type"] == "serial"

        base_match = not filter_text or (
            filter_text in str(row["base_code"]).lower()
            or filter_text in str(row["item_name"]).lower()
            or filter_text in str(row.get("category_name", "")).lower()
        )
        child_any_match = any(
            filter_text in str(d["full_code"]).lower() or filter_text in str(d["size_name"]).lower()
            for d in details
        ) if filter_text else True

        if filter_text and not base_match and not child_any_match:
            return False

        # Есть ли реальный размерный ряд (не только UNI)
        has_sizes = not (len(details) == 1 and details[0]["size_name"].upper() == "UNI")

        top = QTreeWidgetItem(parent) if parent is not None else QTreeWidgetItem(self.tree)
        top.setText(0, row["base_code"])
        top.setText(1, row["item_name"])
        top.setText(2, "")
        top.setText(3, str(row["stock_value"]))
        top.setText(4, row["uom"])
        top.setData(0, Qt.ItemDataRole.UserRole, row["item_id"])
        top.setFlags(top.flags() & ~Qt.ItemFlag.ItemIsEditable)

        if is_serial:
            # Для серийных: уровень 2 — сразу все S/N без промежуточного уровня размеров
            serials = self.db.get_serials_for_item(row["item_id"])
            # Полный код в строке изделия — базовый, т.к. S/N могут быть разных размеров
            top.setText(0, row["base_code"])
            for sn_row in serials:
                sn_node = QTreeWidgetItem(top)
                sn_node.setText(0, sn_row["full_code"])
                sn_node.setText(1, f"S/N: {sn_row['factory_sn']}")
                size = sn_row["size_name"]
                sn_node.setText(2, "" if size.upper() == "UNI" else size)
                sn_node.setText(3, "1")
                sn_node.setText(4, row["uom"])
                sn_node.setFlags(sn_node.flags() & ~Qt.ItemFlag.ItemIsEditable)
                sn_node.setChildIndicatorPolicy(
                    QTreeWidgetItem.ChildIndicatorPolicy.DontShowIndicator
                )
                for col in range(5):
                    sn_node.setForeground(col, QColor("#6B778C"))
            if top.childCount() == 0:
                top.setChildIndicatorPolicy(
                    QTreeWidgetItem.ChildIndicatorPolicy.DontShowIndicator
                )
            top.setExpanded(False)

        elif has_sizes:
            # Мат. средства с размерным рядом
            for d in details:
                if d["stock_value"] == 0:
                    continue
                if filter_text and not base_match:
                    if filter_text not in str(d["full_code"]).lower() \
                            and filter_text not in str(d["size_name"]).lower():
                        continue
                child = QTreeWidgetItem(top)
                child.setText(0, d["full_code"])
                child.setText(1, "")
                child.setText(2, d["size_name"])
                child.setText(3, str(d["stock_value"]))
                child.setText(4, d["uom"])
                child.setFlags(child.flags() & ~Qt.ItemFlag.ItemIsEditable)
                child.setChildIndicatorPolicy(
                    QTreeWidgetItem.ChildIndicatorPolicy.DontShowIndicator
                )
                for col in range(5):
                    child.setForeground(col, QColor("#6B778C"))
            top.setExpanded(False)

        else:
            # Мат. средства без размерного ряда (UNI)
            top.setText(0, details[0]["full_code"] if details else row["base_code"])
            top.setChildIndicatorPolicy(
                QTreeWidgetItem.ChildIndicatorPolicy.DontShowIndicator
            )

        return True

    def reload(self):
        self.tree.clear()
        base_rows = self.db.get_stock_view()

        if self._group_by_category:
            # Группируем по категории
            from collections import OrderedDict
            cat_map: dict = OrderedDict()
            for row in base_rows:
                cat = row["category_name"]
                cat_map.setdefault(cat, []).append(row)

            for cat_name, rows in cat_map.items():
                cat_node = QTreeWidgetItem(self.tree)
                cat_node.setText(0, "")
                cat_node.setText(1, cat_name.upper())
                cat_node.setText(2, "")
                cat_node.setText(3, "")
                cat_node.setText(4, "")
                cat_node.setFlags(cat_node.flags() & ~Qt.ItemFlag.ItemIsEditable)
                font = cat_node.font(1)
                font.setBold(True)
                font.setPointSize(10)
                cat_node.setFont(1, font)
                for col in range(5):
                    cat_node.setBackground(col, QColor("#F4F5F7"))
                    cat_node.setForeground(col, QColor("#6B778C"))

                added = 0
                for row in rows:
                    if self._make_item_node(cat_node, row):
                        added += 1

                cat_node.setExpanded(True)
                # Обновляем счётчик позиций в категории
                cat_node.setText(3, str(added) + " поз.")
        else:
            for row in base_rows:
                self._make_item_node(None, row)

    def on_item_clicked(self, item: QTreeWidgetItem, column: int):
        # Имитация "аккордеона": клик по строке базовой номенклатуры разворачивает/сворачивает размеры
        if item.childCount() > 0:
            item.setExpanded(not item.isExpanded())

    def on_search(self):
        self.reload()


class UnitsTab(QWidget):
    def __init__(self, db: DatabaseManager, units_changed_callback=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.units_changed_callback = units_changed_callback
        # region agent log
        _agent_debug_log(
            hypothesis_id="H1",
            message="UnitsTab.__init__ entered",
            data={},
        )
        # endregion
        self._build_ui()
        self.reload()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(12)

        layout.addWidget(QLabel("Подразделения:"))
        self.list_widget = QListWidget()
        layout.addWidget(self.list_widget, 1)

        btn_layout = QHBoxLayout()
        self.add_btn = QPushButton("Добавить подразделение")
        self.delete_btn = QPushButton("Удалить выбранное")
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.delete_btn)
        layout.addLayout(btn_layout)

        self.add_btn.clicked.connect(self.on_add)
        self.delete_btn.clicked.connect(self.on_delete)

    def reload(self):
        # region agent log
        _agent_debug_log(
            hypothesis_id="H1",
            message="UnitsTab.reload start",
            data={},
        )
        # endregion
        self.list_widget.clear()
        self._units = self.db.get_units()
        for u in self._units:
            text = f"{u['name']}"
            # region agent log
            _agent_debug_log(
                hypothesis_id="H1",
                message="UnitsTab.reload add unit",
                data={"unit_id": u["id"], "name": u["name"]},
            )
            # endregion
            item = QListWidgetItem(text)
            item.setData(Qt.ItemDataRole.UserRole, u["id"])
            self.list_widget.addItem(item)

    def on_add(self):
        name, ok = QInputDialog.getText(self, "Новое подразделение", "Наименование:")
        if not ok:
            return
        name = name.strip()
        if not name:
            QMessageBox.warning(self, "Ошибка", "Наименование не может быть пустым.")
            return
        try:
            self.db.add_unit(name)
        except sqlite3.IntegrityError as e:
            if logger:
                logger.warning("Add unit duplicate: %s", e)
            QMessageBox.warning(self, "Ошибка", "Подразделение с таким именем уже существует.")
            return
        self.reload()
        if self.units_changed_callback:
            self.units_changed_callback()

    def on_delete(self):
        item = self.list_widget.currentItem()
        if not item:
            QMessageBox.warning(self, "Ошибка", "Выберите подразделение для удаления.")
            return
        unit_id = item.data(Qt.ItemDataRole.UserRole)
        name = item.text()
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Удалить подразделение '{name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            self.db.delete_unit(unit_id)
        except sqlite3.IntegrityError as e:
            if logger:
                logger.warning("Delete unit in use: %s", e)
            QMessageBox.warning(
                self,
                "Ошибка",
                "Невозможно удалить подразделение, так как оно уже используется в операциях.",
            )
            return
        self.reload()
        if self.units_changed_callback:
            self.units_changed_callback()


class _NavBtn(QPushButton):
    """Кнопка сайдбара на базе QPushButton.
    Qt рендерит иконку и текст через единый style engine —
    вертикальное выравнивание гарантировано без ручной работы с layout."""

    def __init__(self, text: str, obj_name: str = "NavButton",
                 left_padding: int = 12, parent=None):
        super().__init__(text, parent)
        self.setObjectName(obj_name)
        self.setCheckable(True)
        self.setIconSize(QSize(24, 24))
        self.setFixedHeight(40)
        self.setCursor(Qt.CursorShape.PointingHandCursor)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # region agent log
        _agent_debug_log(
            hypothesis_id="H2",
            message="MainWindow.__init__ entered",
            data={},
        )
        # endregion
        self.setWindowTitle("Складской учет (PyQt6 + SQLite)")
        self.resize(1200, 720)
        self.setMinimumSize(900, 580)

        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), DB_NAME)
        if os.path.exists(db_path):
            _backup_db(db_path)
        self.db = DatabaseManager(db_path)

        self._build_ui()

    def _load_icon(self, filename: str) -> QIcon:
        """Загружает иконку из папки icons/. Возвращает пустой QIcon если файл не найден."""
        base_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(base_dir, "icons", filename)
        if os.path.exists(icon_path):
            return QIcon(icon_path)
        return QIcon()

    def _nav_icon(self, filename: str) -> QIcon:
        """Загружает иконку и возвращает двухрежимный QIcon:
        - Normal  → цвет подписи неактивного пункта (#DEEBFF)
        - Active/Selected → цвет активного пункта (#FFFFFF)
        """
        base_icon = self._load_icon(filename)
        if base_icon.isNull():
            return base_icon

        def colorize(pixmap: QPixmap, hex_color: str) -> QPixmap:
            result = QPixmap(pixmap.size())
            result.fill(QColor(0, 0, 0, 0))
            painter = QPainter(result)
            painter.drawPixmap(0, 0, pixmap)
            painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_SourceIn)
            painter.fillRect(result.rect(), QColor(hex_color))
            painter.end()
            return result

        src = base_icon.pixmap(QSize(24, 24))
        icon = QIcon()
        icon.addPixmap(colorize(src, "#DEEBFF"), QIcon.Mode.Normal,   QIcon.State.Off)
        icon.addPixmap(colorize(src, "#FFFFFF"), QIcon.Mode.Active,   QIcon.State.Off)
        icon.addPixmap(colorize(src, "#FFFFFF"), QIcon.Mode.Selected, QIcon.State.Off)
        icon.addPixmap(colorize(src, "#FFFFFF"), QIcon.Mode.Normal,   QIcon.State.On)
        return icon

    def _build_ui(self):
        # region agent log
        _agent_debug_log(
            hypothesis_id="H2",
            message="MainWindow._build_ui entered",
            data={},
        )
        # endregion

        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QHBoxLayout(central)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # ── Левая панель навигации ──────────────────────────────────────────
        sidebar = QFrame()
        sidebar.setObjectName("Sidebar")
        sidebar.setFixedWidth(240)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(0)

        # Шапка сайдбара — логотип приложения
        app_header = QFrame()
        app_header.setObjectName("AppHeader")
        app_header.setFixedHeight(76)
        app_header_layout = QHBoxLayout(app_header)
        app_header_layout.setContentsMargins(16, 0, 16, 0)
        app_header_layout.setSpacing(10)

        logo_badge = QLabel("WH")
        logo_badge.setObjectName("LogoBadge")
        logo_badge.setFixedSize(32, 32)
        logo_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)

        app_name = QLabel("Склад · ЛТО")
        app_name.setObjectName("AppName")

        app_header_layout.addWidget(logo_badge)
        app_header_layout.addWidget(app_name)
        app_header_layout.addStretch()
        sidebar_layout.addWidget(app_header)

        # Разделитель
        sep = QFrame()
        sep.setObjectName("SidebarDivider")
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setFixedHeight(1)
        sidebar_layout.addWidget(sep)

        # Навигация
        nav_scroll = QWidget()
        nav_scroll.setObjectName("NavArea")
        nav_layout = QVBoxLayout(nav_scroll)
        nav_layout.setContentsMargins(8, 12, 8, 12)
        nav_layout.setSpacing(2)

        # ─ Секция: Склад
        lbl_main = QLabel("СКЛАД")
        lbl_main.setObjectName("NavSection")
        nav_layout.addWidget(lbl_main)

        self.btn_stock = _NavBtn("Остатки на складе", "NavButton")
        self.btn_stock.setIcon(self._nav_icon("ph-warehouse-light.svg"))
        self.btn_stock.setFixedHeight(40)
        nav_layout.addWidget(self.btn_stock)

        nav_layout.addSpacing(8)

        # ─ Секция: Операции
        lbl_ops = QLabel("ОПЕРАЦИИ")
        lbl_ops.setObjectName("NavSection")
        nav_layout.addWidget(lbl_ops)

        self._operations_expanded = True
        self.btn_operations_parent = _NavBtn("Операции  ▼", "NavParent")
        self.btn_operations_parent.setIcon(self._nav_icon("ph-arrows-left-right-light.svg"))
        self.btn_operations_parent.setFixedHeight(40)
        self.btn_operations_parent.setCheckable(False)   # только разворачивает меню
        nav_layout.addWidget(self.btn_operations_parent)

        self.btn_journal = _NavBtn("Журнал операций", "NavChild", left_padding=36)
        self.btn_journal.setFixedHeight(34)
        nav_layout.addWidget(self.btn_journal)

        self.btn_conduct = _NavBtn("Проведение", "NavChild", left_padding=36)
        self.btn_conduct.setFixedHeight(34)
        nav_layout.addWidget(self.btn_conduct)

        nav_layout.addSpacing(8)

        # ─ Секция: Справочники
        lbl_ref = QLabel("СПРАВОЧНИКИ")
        lbl_ref.setObjectName("NavSection")
        nav_layout.addWidget(lbl_ref)

        self.btn_nomenclature = _NavBtn("Номенклатор", "NavButton")
        self.btn_nomenclature.setIcon(self._nav_icon("ph-squares-four-light.svg"))
        self.btn_nomenclature.setFixedHeight(40)
        nav_layout.addWidget(self.btn_nomenclature)

        self.btn_units = _NavBtn("Подразделения", "NavButton")
        self.btn_units.setIcon(self._nav_icon("ph-buildings-light.svg"))
        self.btn_units.setFixedHeight(40)
        nav_layout.addWidget(self.btn_units)

        nav_layout.addStretch()
        sidebar_layout.addWidget(nav_scroll, 1)

        # Подвал сайдбара
        sidebar_footer = QFrame()
        sidebar_footer.setObjectName("SidebarFooterFrame")
        footer_layout = QHBoxLayout(sidebar_footer)
        footer_layout.setContentsMargins(16, 10, 16, 10)
        footer_label = QLabel("PyQt6 · SQLite")
        footer_label.setObjectName("SidebarFooter")
        footer_layout.addWidget(footer_label)
        sidebar_layout.addWidget(sidebar_footer)

        # _NavBtn checkable по умолчанию, дополнительных настроек не требует

        # ── Правая область ─────────────────────────────────────────────────
        right_panel = QFrame()
        right_panel.setObjectName("RightPanel")
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(0)

        # Верхний хедер страницы
        self.page_header = QFrame()
        self.page_header.setObjectName("PageHeader")
        self.page_header.setFixedHeight(76)
        page_header_layout = QHBoxLayout(self.page_header)
        page_header_layout.setContentsMargins(24, 0, 24, 0)
        self.page_title = QLabel("Остатки на складе")
        self.page_title.setObjectName("PageTitle")
        self.page_breadcrumb = QLabel("Склад / Остатки")
        self.page_breadcrumb.setObjectName("PageBreadcrumb")
        title_col = QVBoxLayout()
        title_col.setSpacing(1)
        title_col.setContentsMargins(0, 0, 0, 0)
        title_col.addStretch()
        title_col.addWidget(self.page_breadcrumb)
        title_col.addWidget(self.page_title)
        title_col.addStretch()
        page_header_layout.addLayout(title_col)
        page_header_layout.addStretch()
        right_layout.addWidget(self.page_header)

        # Контент — табы
        self.tabs = QTabWidget()
        self.tabs.setObjectName("MainTabs")
        self.tabs.tabBar().hide()

        self.nomenclature_tab = NomenclatureTab(self.db, self)
        self.journal_tab = JournalTab(self.db, self)
        self.stock_tab = StockTab(self.db, self)
        self.operations_tab = OperationsTab(
            self.db,
            stock_tab_updater=self.stock_tab.reload,
            refresh_journal=self.journal_tab.load_journal,
            parent=self,
        )
        self.units_tab = UnitsTab(
            self.db,
            units_changed_callback=lambda: (
                self.operations_tab.load_units(),
                self.journal_tab.load_units(),
            ),
            parent=self,
        )

        self.tabs.addTab(self.stock_tab, "Склад")
        self.tabs.addTab(self.journal_tab, "Журнал операций")
        self.tabs.addTab(self.operations_tab, "Проведение операций")
        self.tabs.addTab(self.nomenclature_tab, "Номенклатура")
        self.tabs.addTab(self.units_tab, "Подразделения")

        right_layout.addWidget(self.tabs, 1)

        root_layout.addWidget(sidebar)
        root_layout.addWidget(right_panel, 1)

        # Подключаем сигналы
        self.btn_operations_parent.clicked.connect(self._toggle_operations_menu)
        self.btn_stock.clicked.connect(lambda: self._switch_page(0, self.btn_stock, "Остатки на складе", "Склад / Остатки"))
        self.btn_journal.clicked.connect(lambda: self._switch_page(1, self.btn_journal, "Журнал операций", "Операции / Журнал"))
        self.btn_conduct.clicked.connect(lambda: self._switch_page(2, self.btn_conduct, "Проведение операций", "Операции / Провести"))
        self.btn_nomenclature.clicked.connect(lambda: self._switch_page(3, self.btn_nomenclature, "Номенклатор", "Справочники / Номенклатура"))
        self.btn_units.clicked.connect(lambda: self._switch_page(4, self.btn_units, "Подразделения", "Справочники / Подразделения"))

        self._switch_page(0, self.btn_stock, "Остатки на складе", "Склад / Остатки")

    def _toggle_operations_menu(self):
        self._operations_expanded = not self._operations_expanded
        self.btn_journal.setVisible(self._operations_expanded)
        self.btn_conduct.setVisible(self._operations_expanded)
        self.btn_operations_parent.setText(
            "Операции  ▼" if self._operations_expanded else "Операции  ▶"
        )

    def _switch_page(self, index: int, active_btn: "_NavBtn",
                     title: str = "", breadcrumb: str = ""):
        self.tabs.setCurrentIndex(index)
        if title:
            self.page_title.setText(title)
        if breadcrumb:
            self.page_breadcrumb.setText(breadcrumb)
        for btn in (self.btn_stock, self.btn_journal, self.btn_conduct,
                    self.btn_nomenclature, self.btn_units):
            btn.setChecked(btn is active_btn)


def main():
    global logger
    logger = _setup_logging()

    app = QApplication(sys.argv)

    # ── Загрузка шрифта Open Sans ────────────────────────────────────────────
    _font_family = "Open Sans"
    _fonts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
    _font_files = [
        "OpenSans-Regular.ttf",
        "OpenSans-SemiBold.ttf",
        "OpenSans-Bold.ttf",
        "OpenSans-Italic.ttf",
    ]
    _loaded = False
    for _fname in _font_files:
        _fpath = os.path.join(_fonts_dir, _fname)
        if os.path.exists(_fpath):
            QFontDatabase.addApplicationFont(_fpath)
            _loaded = True

    if _loaded:
        _base_font = QFont("Open Sans", 10)
        app.setFont(_base_font)
    else:
        _font_family = '-apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif'

    # ── Atlassian Design System ──────────────────────────────────────────────
    app.setStyleSheet(
        """
        /* ─ BASE ─────────────────────────────────────────────────────────── */
        QWidget {
            background-color: #F4F5F7;
            font-family: "Open Sans", -apple-system, "Segoe UI", Roboto, Arial, sans-serif;
            font-size: 13px;
            color: #172B4D;
        }

        /* ─ SIDEBAR ──────────────────────────────────────────────────────── */
        QFrame#Sidebar {
            background-color: #0747A6;
            border: none;
        }

        QFrame#AppHeader {
            background-color: #0052CC;
        }

        QLabel#LogoBadge {
            background-color: #FFFFFF;
            color: #0052CC;
            font-weight: 700;
            font-size: 13px;
            border-radius: 4px;
        }

        QLabel#AppName {
            color: #FFFFFF;
            font-weight: 600;
            font-size: 14px;
        }

        QFrame#SidebarDivider {
            background-color: #0052CC;
            border: none;
        }

        QWidget#NavArea {
            background-color: #0747A6;
        }

        QLabel#NavSection {
            color: #7AB2F4;
            font-size: 10px;
            font-weight: 700;
            padding: 6px 4px 2px 4px;
            background-color: transparent;
        }

        QPushButton#NavButton,
        QPushButton#NavParent {
            background: transparent;
            border: none;
            border-radius: 3px;
            color: #DEEBFF;
            font-size: 13px;
            font-weight: 400;
            text-align: left;
            padding-left: 12px;
            padding-right: 12px;
        }

        QPushButton#NavParent {
            font-weight: 600;
        }

        QPushButton#NavChild {
            background: transparent;
            border: none;
            border-radius: 3px;
            color: #B3D4FF;
            font-size: 12px;
            text-align: left;
            padding-left: 36px;
            padding-right: 12px;
        }

        QPushButton#NavButton:hover,
        QPushButton#NavParent:hover,
        QPushButton#NavChild:hover {
            background: rgba(255, 255, 255, 25);
            color: #FFFFFF;
        }

        QPushButton#NavButton:checked,
        QPushButton#NavParent:checked,
        QPushButton#NavChild:checked {
            background: #0052CC;
            color: #FFFFFF;
            font-weight: 600;
        }

        QFrame#SidebarFooterFrame {
            background-color: #043584;
            border: none;
        }

        QLabel#SidebarFooter {
            color: #7AB2F4;
            font-size: 11px;
            background-color: transparent;
        }

        /* ─ RIGHT PANEL ──────────────────────────────────────────────────── */
        QFrame#RightPanel {
            background-color: #F4F5F7;
        }

        QFrame#PageHeader {
            background-color: #FFFFFF;
            border-bottom: 2px solid #DFE1E6;
        }

        QLabel#PageTitle {
            font-size: 20px;
            font-weight: 600;
            color: #172B4D;
            background-color: transparent;
        }

        QLabel#PageBreadcrumb {
            font-size: 11px;
            color: #6B778C;
            background-color: transparent;
        }

        QTabWidget#MainTabs::pane {
            border: none;
            background-color: #F4F5F7;
        }

        /* ─ INPUTS ───────────────────────────────────────────────────────── */
        QLineEdit,
        QComboBox,
        QSpinBox,
        QDateEdit {
            background-color: #FFFFFF;
            border: 2px solid #DFE1E6;
            border-radius: 3px;
            padding: 0px 8px;
            color: #172B4D;
            min-height: 32px;
            max-height: 32px;
            selection-background-color: #DEEBFF;
            selection-color: #172B4D;
        }

        QLineEdit:focus,
        QComboBox:focus,
        QSpinBox:focus,
        QDateEdit:focus {
            border-color: #4C9AFF;
            outline: none;
        }

        QLineEdit:hover,
        QComboBox:hover,
        QSpinBox:hover,
        QDateEdit:hover {
            border-color: #B3BAC5;
        }

        QComboBox::drop-down {
            border: none;
            padding-right: 8px;
        }

        QComboBox QAbstractItemView {
            background-color: #FFFFFF;
            border: 1px solid #DFE1E6;
            selection-background-color: #DEEBFF;
            selection-color: #172B4D;
            outline: none;
        }

        /* ─ BUTTONS ──────────────────────────────────────────────────────── */
        QPushButton {
            background-color: #0052CC;
            color: #FFFFFF;
            border: none;
            border-radius: 3px;
            padding: 0px 12px;
            min-height: 32px;
            max-height: 32px;
            font-weight: 500;
            font-size: 14px;
        }

        QPushButton:hover {
            background-color: #0747A6;
        }

        QPushButton:pressed {
            background-color: #043584;
        }

        QPushButton:disabled {
            background-color: #DFE1E6;
            color: #97A0AF;
        }

        QPushButton:checked {
            background-color: #043584;
            color: #FFFFFF;
            border: 2px solid #4C9AFF;
        }

        QPushButton#DangerBtn {
            background-color: transparent;
            color: #DE350B;
            border: 1.5px solid #DE350B;
            border-radius: 3px;
        }

        QPushButton#DangerBtn:hover {
            background-color: #FFEBE6;
        }

        QPushButton#DangerBtn:pressed {
            background-color: #FFBDAD;
        }

        QPushButton#ExportBtn {
            background-color: transparent;
            border: 1px solid #B3BAC5;
            border-radius: 3px;
            padding: 0px;
        }

        QPushButton#ExportBtn:hover {
            background-color: #F4F5F7;
            border-color: #97A0AF;
        }

        QPushButton#ExportBtn:pressed {
            background-color: #DFE1E6;
        }

        /* ─ RADIO ────────────────────────────────────────────────────────── */
        QRadioButton {
            color: #172B4D;
            spacing: 8px;
            background-color: transparent;
        }

        QRadioButton::indicator {
            width: 16px;
            height: 16px;
            border-radius: 8px;
            border: 2px solid #6B778C;
            background-color: #FFFFFF;
        }

        QRadioButton::indicator:hover {
            border-color: #0052CC;
        }

        QRadioButton::indicator:checked {
            border-color: #0052CC;
            background-color: #0052CC;
        }

        /* ─ LABELS ───────────────────────────────────────────────────────── */
        QLabel {
            color: #172B4D;
            background-color: transparent;
        }

        /* ─ LIST ─────────────────────────────────────────────────────────── */
        QListWidget {
            background-color: #FFFFFF;
            border: 1px solid #DFE1E6;
            border-radius: 3px;
            outline: none;
            color: #172B4D;
        }

        QListWidget::item {
            min-height: 36px;
            padding: 4px 8px;
            border-bottom: 1px solid #F4F5F7;
            color: #172B4D;
        }

        QListWidget::item:selected {
            background-color: #DEEBFF;
            color: #172B4D;
        }

        QListWidget::item:hover {
            background-color: #F4F5F7;
        }

        /* ─ TABLE ────────────────────────────────────────────────────────── */
        QTableWidget {
            background-color: #FFFFFF;
            border: 1px solid #DFE1E6;
            border-radius: 3px;
            gridline-color: #F4F5F7;
            outline: none;
            color: #172B4D;
        }

        QTableWidget::item {
            padding: 6px 8px;
            border-bottom: 1px solid #F4F5F7;
            color: #172B4D;
        }

        QTableWidget::item:selected {
            background-color: #DEEBFF;
            color: #172B4D;
        }

        QTableWidget::item:hover {
            background-color: #F4F5F7;
        }

        QHeaderView::section {
            background-color: #F4F5F7;
            border: none;
            border-bottom: 2px solid #DFE1E6;
            border-right: 1px solid #DFE1E6;
            padding: 8px 8px;
            font-weight: 600;
            font-size: 11px;
            color: #6B778C;
            text-transform: uppercase;
        }

        /* ─ TREE ─────────────────────────────────────────────────────────── */
        QTreeWidget {
            background-color: #FFFFFF;
            border: 1px solid #DFE1E6;
            border-radius: 3px;
            gridline-color: #F4F5F7;
            outline: none;
            color: #172B4D;
        }

        QTreeView::item {
            padding: 6px 8px;
            border-bottom: 1px solid #F4F5F7;
            color: #172B4D;
            min-height: 0;
        }

        QTreeView::item:focus {
            outline: none;
            border: none;
            border-bottom: 1px solid #F4F5F7;
        }

        QTreeView::item:selected {
            background-color: #DEEBFF;
            color: #172B4D;
        }

        QTreeView::item:hover {
            background-color: #F4F5F7;
        }

        QTreeView::branch {
            background-color: #FFFFFF;
            border-bottom: 1px solid #F4F5F7;
        }


        /* ─ SCROLLBAR ────────────────────────────────────────────────────── */
        QScrollBar:vertical {
            width: 8px;
            background-color: #F4F5F7;
            border: none;
        }

        QScrollBar::handle:vertical {
            background-color: #C1C7D0;
            border-radius: 4px;
            min-height: 30px;
        }

        QScrollBar::handle:vertical:hover {
            background-color: #97A0AF;
        }

        QScrollBar::add-line:vertical,
        QScrollBar::sub-line:vertical {
            height: 0px;
        }

        QScrollBar:horizontal {
            height: 8px;
            background-color: #F4F5F7;
            border: none;
        }

        QScrollBar::handle:horizontal {
            background-color: #C1C7D0;
            border-radius: 4px;
            min-width: 30px;
        }

        QScrollBar::handle:horizontal:hover {
            background-color: #97A0AF;
        }

        QScrollBar::add-line:horizontal,
        QScrollBar::sub-line:horizontal {
            width: 0px;
        }

        /* ─ CALENDAR POPUP ───────────────────────────────────────────────── */
        QCalendarWidget {
            background-color: #FFFFFF;
            border: 1px solid #DFE1E6;
            border-radius: 4px;
        }

        QCalendarWidget QWidget#qt_calendar_navigationbar {
            background-color: #0052CC;
            padding: 4px 8px;
            border-radius: 4px 4px 0 0;
        }

        QCalendarWidget QToolButton {
            background-color: transparent;
            color: #FFFFFF;
            border: none;
            border-radius: 3px;
            font-size: 13px;
            font-weight: 600;
            padding: 4px 8px;
            min-height: 28px;
        }

        QCalendarWidget QToolButton:hover {
            background-color: rgba(255, 255, 255, 30);
        }

        QCalendarWidget QToolButton::menu-indicator {
            image: none;
        }

        QCalendarWidget QToolButton#qt_calendar_prevmonth,
        QCalendarWidget QToolButton#qt_calendar_nextmonth {
            padding: 2px 10px;
            min-width: 28px;
        }

        QCalendarWidget QSpinBox {
            background-color: transparent;
            color: #FFFFFF;
            border: none;
            font-size: 13px;
            font-weight: 600;
            min-height: 28px;
            max-height: 28px;
            selection-background-color: rgba(255,255,255,50);
            selection-color: #FFFFFF;
        }

        QCalendarWidget QSpinBox::up-button,
        QCalendarWidget QSpinBox::down-button {
            width: 0px;
            height: 0px;
        }

        QCalendarWidget QAbstractItemView {
            background-color: #FFFFFF;
            selection-background-color: #0052CC;
            selection-color: #FFFFFF;
            outline: none;
            font-size: 12px;
            gridline-color: transparent;
        }

        QCalendarWidget QAbstractItemView:enabled {
            color: #172B4D;
        }

        QCalendarWidget QAbstractItemView:disabled {
            color: #C1C7D0;
        }

        QCalendarWidget QAbstractItemView::item {
            border-radius: 3px;
            padding: 3px;
            min-width: 28px;
            min-height: 28px;
        }

        QCalendarWidget QAbstractItemView::item:hover {
            background-color: #DEEBFF;
            color: #0052CC;
        }

        QCalendarWidget QAbstractItemView::item:selected {
            background-color: #0052CC;
            color: #FFFFFF;
            font-weight: 600;
        }

        QCalendarWidget QWidget {
            alternate-background-color: #F4F5F7;
        }

        QCalendarWidget QLabel {
            color: #6B778C;
            font-size: 11px;
            font-weight: 700;
            background: transparent;
        }
        """
    )

    window = MainWindow()
    # region agent log
    _agent_debug_log(
        hypothesis_id="H2",
        message="Main window created",
        data={},
    )
    # endregion
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

