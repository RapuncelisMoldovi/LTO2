import os
import logging
import sqlite3
from datetime import datetime
from itertools import groupby

logger = logging.getLogger("LTO.database")

# Категории номенклатуры (отображаются в UI как есть).
NOMENCLATURE_CATEGORY_FLIGHT = "лётное снаряжение"
NOMENCLATURE_CATEGORY_PARACHUTE = "парашютно-десантное"
NOMENCLATURE_CATEGORIES: tuple[str, ...] = (
    NOMENCLATURE_CATEGORY_FLIGHT,
    NOMENCLATURE_CATEGORY_PARACHUTE,
)


def normalize_nomenclature_category(value: str | None) -> str:
    """Возвращает допустимую категорию или категорию по умолчанию."""
    s = (value or "").strip()
    if s in NOMENCLATURE_CATEGORIES:
        return s
    return NOMENCLATURE_CATEGORY_FLIGHT


def _sqlite_py_lower(value: str | bytes | None) -> str:
    """Регистрируется в SQLite: нижний регистр через Python (кириллица и пр.). Встроенный lower() в SQLite ASCII-only."""
    if value is None:
        return ""
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8", errors="replace").lower()
        except Exception:
            return ""
    return str(value).lower()


class DatabaseManager:
    def __init__(self, path: str):
        self.path = path
        try:
            self.conn = sqlite3.connect(self.path)
            self.conn.row_factory = sqlite3.Row
            self.conn.create_function("py_lower", 1, _sqlite_py_lower)
            self._init_db()
            self.conn.execute("PRAGMA foreign_keys = ON")
        except Exception as e:
            logger.exception("Database init failed: %s", e)
            raise

    def close(self):
        if self.conn:
            self.conn.close()
            self.conn = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False

    def _init_db(self):
        cur = self.conn.cursor()

        # Подразделения
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS units (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
            """
        )

        # Изделия
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                base_code TEXT NOT NULL,
                uom TEXT NOT NULL,
                type TEXT NOT NULL CHECK (type IN ('qty', 'serial'))
            )
            """
        )

        # Варианты (размеры)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS variants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_id INTEGER NOT NULL REFERENCES items(id) ON DELETE CASCADE,
                size_name TEXT NOT NULL,
                full_code TEXT NOT NULL UNIQUE
            )
            """
        )

        # Количественный учет (материальные средства)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS stock_qty (
                variant_id INTEGER PRIMARY KEY REFERENCES variants(id) ON DELETE CASCADE,
                quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0)
            )
            """
        )

        # Серийный учет (основные средства)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS stock_serial (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                variant_id INTEGER NOT NULL REFERENCES variants(id) ON DELETE CASCADE,
                factory_sn TEXT NOT NULL UNIQUE,
                manufacture_year INTEGER
            )
            """
        )

        # Журнал операций
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS journal (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                op_type TEXT NOT NULL CHECK (op_type IN ('IN','OUT')),
                variant_id INTEGER NOT NULL REFERENCES variants(id),
                quantity INTEGER,
                factory_sn TEXT,
                manufacture_year INTEGER,
                unit_id INTEGER REFERENCES units(id),
                doc_name TEXT
            )
            """
        )

        # Наряды (заявки подразделений)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS work_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                order_no TEXT NOT NULL,
                unit_id INTEGER REFERENCES units(id),
                description TEXT,
                status TEXT NOT NULL CHECK (status IN ('не реализован', 'реализован частично', 'реализован'))
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS work_order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                work_order_id INTEGER NOT NULL REFERENCES work_orders(id) ON DELETE CASCADE,
                variant_id INTEGER NOT NULL REFERENCES variants(id),
                quantity INTEGER NOT NULL CHECK (quantity > 0)
            )
            """
        )

        # Audit log для критичных операций (приоритет 2)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                entity_id INTEGER,
                details TEXT
            )
            """
        )

        # Подразделения по умолчанию — только при самом первом запуске (новая БД: нет ни units, ни items).
        # Флаг _init_flags не даёт восстанавливать их после того, как пользователь удалил подразделения.
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS _init_flags (
                name TEXT PRIMARY KEY,
                value INTEGER NOT NULL
            )
            """
        )
        cur.execute("SELECT COUNT(*) FROM units")
        units_count = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM items")
        items_count = cur.fetchone()[0]
        cur.execute("SELECT value FROM _init_flags WHERE name = 'default_units_added'")
        default_units_done = cur.fetchone() is not None
        if units_count == 0 and items_count == 0 and not default_units_done:
            cur.execute("INSERT INTO units(name) VALUES (?)", ("Склад-основной",))
            cur.execute("INSERT INTO units(name) VALUES (?)", ("Цех-1",))
            cur.execute(
                "INSERT OR REPLACE INTO _init_flags (name, value) VALUES ('default_units_added', 1)"
            )

        self.conn.commit()

        # Миграции: добавляем поля в journal, если их нет (старые БД)
        cur.execute("PRAGMA table_info(journal)")
        cols = [r[1] for r in cur.fetchall()]
        if "doc_name" not in cols:
            cur.execute("ALTER TABLE journal ADD COLUMN doc_name TEXT")
        if "work_order_id" not in cols:
            cur.execute("ALTER TABLE journal ADD COLUMN work_order_id INTEGER REFERENCES work_orders(id)")
        if "manufacture_year" not in cols:
            cur.execute("ALTER TABLE journal ADD COLUMN manufacture_year INTEGER")
        self.conn.commit()

        cur.execute("PRAGMA table_info(stock_serial)")
        ss_cols = [r[1] for r in cur.fetchall()]
        if "manufacture_year" not in ss_cols:
            cur.execute("ALTER TABLE stock_serial ADD COLUMN manufacture_year INTEGER")
        self.conn.commit()

        # Категория номенклатуры (лётное / парашютно-десантное и др.)
        cur.execute("PRAGMA table_info(items)")
        item_cols = [r[1] for r in cur.fetchall()]
        if "category" not in item_cols:
            cur.execute(
                "ALTER TABLE items ADD COLUMN category TEXT NOT NULL DEFAULT 'лётное снаряжение'"
            )
        self.conn.commit()

        # Миграция: переименование статусов work_orders (ж → м род)
        cur.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='work_orders'")
        create_sql = (cur.fetchone() or [None])[0] or ""
        if "реализована" in create_sql:
            cur.execute("ALTER TABLE work_orders RENAME TO _work_orders_old")
            cur.execute(
                """
                CREATE TABLE work_orders (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT NOT NULL,
                    order_no TEXT NOT NULL,
                    unit_id INTEGER REFERENCES units(id),
                    description TEXT,
                    status TEXT NOT NULL CHECK (status IN ('не реализован', 'реализован частично', 'реализован'))
                )
                """
            )
            cur.execute(
                """
                INSERT INTO work_orders (id, created_at, order_no, unit_id, description, status)
                SELECT id, created_at, order_no, unit_id, description,
                    CASE status
                        WHEN 'не реализована' THEN 'не реализован'
                        WHEN 'реализована частично' THEN 'реализован частично'
                        WHEN 'реализована' THEN 'реализован'
                        ELSE status
                    END
                FROM _work_orders_old
                """
            )
            cur.execute("DROP TABLE _work_orders_old")
            self.conn.commit()

        # Repair: if a previous run had PRAGMA foreign_keys=ON during the RENAME
        # migration above, SQLite rewrote FK references in child tables to point
        # at _work_orders_old. Fix them by recreating the affected tables.
        for tbl in ("journal", "work_order_items"):
            row = cur.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (tbl,)
            ).fetchone()
            if row and row[0] and "_work_orders_old" in row[0]:
                logger.warning("Repairing corrupted FK in table %s", tbl)
                cur.execute(f"ALTER TABLE {tbl} RENAME TO _{tbl}_repair")
                fixed_sql = row[0].replace("_work_orders_old", "work_orders")
                cur.execute(fixed_sql)
                cur.execute(f"INSERT INTO {tbl} SELECT * FROM _{tbl}_repair")
                cur.execute(f"DROP TABLE _{tbl}_repair")
                self.conn.commit()

        self._ensure_indices()
        logger.info("Database initialized: %s", self.path)

    def _ensure_indices(self):
        cur = self.conn.cursor()
        cur.execute("CREATE INDEX IF NOT EXISTS idx_journal_variant ON journal(variant_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_journal_doc_name ON journal(doc_name)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_journal_date ON journal(date)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_journal_work_order ON journal(work_order_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_variants_item ON variants(item_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_items_category ON items(category)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_stock_serial_variant ON stock_serial(variant_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_work_order_items_wo ON work_order_items(work_order_id)")
        self.conn.commit()

    # --- Items & Variants ---

    def get_items(self, category: str | None = None):
        cur = self.conn.cursor()
        if category:
            cur.execute(
                """
                SELECT i.id, i.name, i.base_code, i.uom, i.type, i.category
                FROM items i
                WHERE i.category = ?
                ORDER BY i.name COLLATE NOCASE
                """,
                (normalize_nomenclature_category(category),),
            )
        else:
            cur.execute(
                """
                SELECT i.id, i.name, i.base_code, i.uom, i.type, i.category
                FROM items i
                ORDER BY i.name COLLATE NOCASE
                """
            )
        return cur.fetchall()

    def add_item(
        self,
        name: str,
        base_code: str,
        uom: str,
        item_type: str,
        category: str | None = None,
    ):
        cat = normalize_nomenclature_category(category)
        try:
            self.conn.execute("BEGIN")
            cur = self.conn.cursor()
            cur.execute(
                "INSERT INTO items(name, base_code, uom, type, category) VALUES (?, ?, ?, ?, ?)",
                (name.strip(), base_code.strip(), uom.strip(), item_type, cat),
            )
            iid = cur.lastrowid
            self._audit(
                "ITEM_INSERT",
                "items",
                iid,
                f"name={name.strip()} base_code={base_code.strip()} category={cat}",
                commit=False,
            )
            self.conn.commit()
            return iid
        except Exception:
            self.conn.rollback()
            raise

    def get_variants_for_item(self, item_id: int):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT id, size_name, full_code
            FROM variants
            WHERE item_id = ?
            ORDER BY size_name COLLATE NOCASE
            """,
            (item_id,),
        )
        return cur.fetchall()

    def get_nomenclature_tree_data(self, category: str | None = None) -> list[dict]:
        """Изделия и все варианты одним запросом (без N+1 к get_variants_for_item)."""
        cur = self.conn.cursor()
        cat = normalize_nomenclature_category(category) if category else None
        if cat:
            cur.execute(
                """
                SELECT
                    i.id AS item_id,
                    i.name AS item_name,
                    i.base_code,
                    i.uom,
                    i.type AS item_type,
                    v.id AS variant_id,
                    v.size_name,
                    v.full_code
                FROM items i
                LEFT JOIN variants v ON v.item_id = i.id
                WHERE i.category = ?
                ORDER BY i.name COLLATE NOCASE, v.size_name COLLATE NOCASE
                """,
                (cat,),
            )
        else:
            cur.execute(
                """
                SELECT
                    i.id AS item_id,
                    i.name AS item_name,
                    i.base_code,
                    i.uom,
                    i.type AS item_type,
                    v.id AS variant_id,
                    v.size_name,
                    v.full_code
                FROM items i
                LEFT JOIN variants v ON v.item_id = i.id
                ORDER BY i.name COLLATE NOCASE, v.size_name COLLATE NOCASE
                """
            )
        rows = cur.fetchall()
        out: list[dict] = []
        current: dict | None = None
        for row in rows:
            iid = row["item_id"]
            if current is None or current["id"] != iid:
                if current is not None:
                    out.append(current)
                current = {
                    "id": iid,
                    "name": row["item_name"],
                    "base_code": row["base_code"],
                    "uom": row["uom"],
                    "type": row["item_type"],
                    "variants": [],
                }
            vid = row["variant_id"]
            if vid is not None:
                current["variants"].append(
                    {
                        "id": vid,
                        "size_name": row["size_name"],
                        "full_code": row["full_code"],
                    }
                )
        if current is not None:
            out.append(current)
        return out

    def get_item(self, item_id: int):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM items WHERE id = ?", (item_id,))
        return cur.fetchone()

    def add_variant(self, item_id: int, size_name: str, full_code: str, item_type: str):
        try:
            self.conn.execute("BEGIN")
            cur = self.conn.cursor()
            cur.execute(
                "INSERT INTO variants(item_id, size_name, full_code) VALUES (?, ?, ?)",
                (item_id, size_name.strip(), full_code.strip()),
            )
            variant_id = cur.lastrowid
            if item_type == "qty":
                cur.execute(
                    "INSERT OR IGNORE INTO stock_qty(variant_id, quantity) VALUES (?, 0)",
                    (variant_id,),
                )
            self._audit("VARIANT_INSERT", "variants", variant_id, f"item_id={item_id} size={size_name.strip()} full_code={full_code.strip()}", commit=False)
            self.conn.commit()
            return variant_id
        except Exception:
            self.conn.rollback()
            raise

    # Во второй колонке Excel: строка с одним из этих значений задаёт базовый н/н изделия (колонка C); остальные строки с тем же наименованием — размеры.
    EXCEL_BASE_ROW_MARKERS = ("н/н (базовый)", "Без размера")

    def import_nomenclature_from_excel(
        self,
        file_path: str,
        *,
        skip_header_row: bool = True,
        category: str | None = None,
    ) -> tuple[int, int, list[str]]:
        """
        Импорт номенклатуры из Excel.
        Колонки: A — наименование, B — размер/маркер, C — н/н, D — ед. изм., E — тип учёта.
        E пусто — материальные средства (qty), без заводского номера при поступлении/выдаче.
        E = «sn» — основные средства (serial), требуется заводской номер при поступлении и выдаче.
        """
        import_category = normalize_nomenclature_category(category)
        try:
            import openpyxl
        except ImportError:
            return 0, 0, ["Установите openpyxl: pip install openpyxl"]

        errors: list[str] = []
        items_added = 0
        variants_added = 0
        markers = tuple(m.strip() for m in self.EXCEL_BASE_ROW_MARKERS)

        def _is_base_row(col_b: str | None) -> bool:
            if col_b is None:
                return False
            s = str(col_b).strip()
            return any(s == m or s.lower() == m.lower() for m in markers)

        def _excel_type_to_item_type(col_e: str | None) -> str:
            """Колонка E: пусто → qty (мат. средства), «sn» → serial (основные средства)."""
            if col_e is None:
                return "qty"
            s = str(col_e).strip().lower()
            return "serial" if s == "sn" else "qty"

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                return 0, 0, ["В книге нет активного листа"]

            rows: list[tuple[str, str, str, str, str]] = []
            for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
                if not row or row[0] is None:
                    continue
                name = (row[0] or "").strip()
                if not name:
                    continue
                col_b = (row[1] if len(row) > 1 else None) or ""
                col_b = str(col_b).strip()
                code = (row[2] if len(row) > 2 else None) or ""
                code = str(code).strip()
                uom = (row[3] if len(row) > 3 else None) or ""
                uom = str(uom).strip() or "шт"
                col_e = (row[4] if len(row) > 4 else None) or ""
                rows.append((name, col_b, code, uom, str(col_e).strip()))
            wb.close()

            if skip_header_row and rows:
                first_cell = str(rows[0][0]).strip().lower()
                if first_cell in ("наименование", "название", "name", "номер", "n", "наименование изделия"):
                    rows = rows[1:]
        except Exception as e:
            return 0, 0, [f"Ошибка чтения файла: {e}"]

        if not rows:
            return 0, 0, ["В файле нет данных (ожидаются колонки: наименование, размер/маркер, н/н, ед. изм., тип [sn или пусто])."]

        def _name_key(r):
            return r[0]

        for name, group in groupby(sorted(rows, key=_name_key), key=_name_key):
            group_list = list(group)
            base_row = next((r for r in group_list if _is_base_row(r[1])), None)
            if not base_row:
                errors.append(f"«{name}»: нет строки с маркером «н/н (базовый)» или «Без размера» во второй колонке")
                continue
            base_code = base_row[2].strip() if base_row[2] else ""
            if not base_code:
                errors.append(f"«{name}»: в базовой строке (н/н (базовый) или Без размера) не указан номенклатурный номер (колонка C)")
                continue
            uom = base_row[3] or "шт"
            item_type = _excel_type_to_item_type(base_row[4] if len(base_row) > 4 else None)

            try:
                item_id = self.add_item(
                    name, base_code, uom, item_type, category=import_category
                )
                items_added += 1
            except sqlite3.IntegrityError as e:
                errors.append(f"«{name}»: не удалось добавить изделие — {e}")
                continue

            for row in group_list:
                _name, col_b, code, _uom = row[0], row[1], row[2], row[3]
                if _is_base_row(col_b):
                    size_name = "Без размера"
                    full_code = base_code
                else:
                    size_name = col_b or "Без размера"
                    full_code = code.strip() if code else base_code
                if not full_code:
                    full_code = base_code
                try:
                    self.add_variant(item_id, size_name, full_code, item_type)
                    variants_added += 1
                except sqlite3.IntegrityError:
                    errors.append(f"«{name}», размер «{size_name}»: дубликат н/н «{full_code}» — пропущено")

        return items_added, variants_added, errors

    def update_item(
        self,
        item_id: int,
        name: str,
        base_code: str,
        uom: str,
        item_type: str,
        category: str | None = None,
    ):
        cat = normalize_nomenclature_category(category)
        try:
            self.conn.execute("BEGIN")
            cur = self.conn.cursor()
            cur.execute(
                """
                UPDATE items
                SET name = ?, base_code = ?, uom = ?, type = ?, category = ?
                WHERE id = ?
                """,
                (name.strip(), base_code.strip(), uom.strip(), item_type, cat, item_id),
            )
            self._audit(
                "ITEM_UPDATE",
                "items",
                item_id,
                f"name={name.strip()} base_code={base_code.strip()} category={cat}",
                commit=False,
            )
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def update_variant(self, variant_id: int, size_name: str, full_code: str):
        try:
            self.conn.execute("BEGIN")
            cur = self.conn.cursor()
            cur.execute(
                "UPDATE variants SET size_name = ?, full_code = ? WHERE id = ?",
                (size_name.strip(), full_code.strip(), variant_id),
            )
            self._audit("VARIANT_UPDATE", "variants", variant_id, f"size={size_name.strip()} full_code={full_code.strip()}", commit=False)
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    # --- Units ---

    def get_units(self):
        cur = self.conn.cursor()
        cur.execute("SELECT id, name FROM units ORDER BY name COLLATE NOCASE")
        return cur.fetchall()

    def add_unit(self, name: str):
        try:
            self.conn.execute("BEGIN")
            cur = self.conn.cursor()
            cur.execute("INSERT INTO units(name) VALUES (?)", (name.strip(),))
            uid = cur.lastrowid
            self._audit("UNIT_INSERT", "units", uid, f"name={name.strip()}", commit=False)
            self.conn.commit()
            return uid
        except Exception:
            self.conn.rollback()
            raise

    def delete_unit(self, unit_id: int):
        try:
            self.conn.execute("BEGIN")
            cur = self.conn.cursor()
            cur.execute("DELETE FROM units WHERE id = ?", (unit_id,))
            self._audit("UNIT_DELETE", "units", unit_id, "", commit=False)
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def has_journal_entries_for_item(self, item_id: int) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT COUNT(*) AS cnt FROM journal j
            JOIN variants v ON v.id = j.variant_id
            WHERE v.item_id = ?
            """,
            (item_id,),
        )
        return (cur.fetchone()["cnt"] or 0) > 0

    def has_journal_entries_for_variant(self, variant_id: int) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT COUNT(*) AS cnt FROM journal WHERE variant_id = ?",
            (variant_id,),
        )
        return (cur.fetchone()["cnt"] or 0) > 0

    def work_order_items_count_for_variant(self, variant_id: int) -> int:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT COUNT(*) AS cnt FROM work_order_items WHERE variant_id = ?",
            (variant_id,),
        )
        return int(cur.fetchone()["cnt"] or 0)

    def delete_item(self, item_id: int):
        """Удаляет изделие и все связанные варианты, остатки и записи журнала."""
        try:
            self.conn.execute("BEGIN")
            cur = self.conn.cursor()
            cur.execute("SELECT id FROM variants WHERE item_id = ?", (item_id,))
            variant_ids = [r["id"] for r in cur.fetchall()]
            if variant_ids:
                placeholders = ",".join("?" * len(variant_ids))
                cur.execute(f"DELETE FROM journal WHERE variant_id IN ({placeholders})", variant_ids)
                cur.execute(f"DELETE FROM stock_qty WHERE variant_id IN ({placeholders})", variant_ids)
                cur.execute(f"DELETE FROM stock_serial WHERE variant_id IN ({placeholders})", variant_ids)
            cur.execute("DELETE FROM variants WHERE item_id = ?", (item_id,))
            cur.execute("DELETE FROM items WHERE id = ?", (item_id,))
            self._audit("ITEM_DELETE", "items", item_id, f"variants={variant_ids}", commit=False)
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def delete_variant(self, variant_id: int):
        """Удаляет вариант (размер) и связанные журнал, остатки, S/N и строки нарядов."""
        try:
            self.conn.execute("BEGIN")
            cur = self.conn.cursor()
            cur.execute("DELETE FROM work_order_items WHERE variant_id = ?", (variant_id,))
            cur.execute("DELETE FROM journal WHERE variant_id = ?", (variant_id,))
            cur.execute("DELETE FROM stock_qty WHERE variant_id = ?", (variant_id,))
            cur.execute("DELETE FROM stock_serial WHERE variant_id = ?", (variant_id,))
            cur.execute("DELETE FROM variants WHERE id = ?", (variant_id,))
            self._audit("VARIANT_DELETE", "variants", variant_id, "", commit=False)
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def get_or_create_unit(self, name: str | None):
        if not name:
            return None
        name = name.strip()
        if not name:
            return None
        cur = self.conn.cursor()
        cur.execute("SELECT id FROM units WHERE name = ?", (name,))
        row = cur.fetchone()
        if row:
            return row["id"]
        cur.execute("INSERT INTO units(name) VALUES (?)", (name,))
        self.conn.commit()
        return cur.lastrowid

    # --- Work Orders (Наряды) ---

    def get_work_orders_brief(self):
        """Lightweight list for combo boxes — no status recomputation."""
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT wo.id, wo.order_no, wo.unit_id, u.name AS unit_name, wo.status
            FROM work_orders wo
            LEFT JOIN units u ON u.id = wo.unit_id
            ORDER BY wo.id DESC
            """
        )
        return cur.fetchall()

    def get_work_order_fulfillment_pct(self, work_order_id: int) -> int:
        """Returns fulfillment percentage (0–100) for a work order."""
        items = self.get_work_order_items(work_order_id)
        if not items:
            return 0
        issued = self.get_work_order_item_issue_stats(work_order_id)
        total_req = sum(int(r["requested_qty"] or 0) for r in items)
        total_done = sum(int(issued.get(r["variant_id"], 0)) for r in items)
        return round(total_done / total_req * 100) if total_req > 0 else 0

    def get_work_order_remaining_items(self, work_order_id: int):
        """Returns work order items with remaining (not yet issued) quantities."""
        items = self.get_work_order_items(work_order_id)
        issued = self.get_work_order_item_issue_stats(work_order_id)
        result = []
        for row in items:
            req = int(row["requested_qty"] or 0)
            done = int(issued.get(row["variant_id"], 0))
            remaining = max(0, req - done)
            if remaining > 0:
                result.append({
                    "variant_id": row["variant_id"],
                    "item_name": row["item_name"],
                    "full_code": row["full_code"],
                    "size_name": row["size_name"],
                    "item_type": row["item_type"],
                    "qty": remaining,
                    "sn": None,
                })
        return result

    def get_work_orders(self, search_text: str = ""):
        cur = self.conn.cursor()
        pattern = f"%{search_text.strip().lower()}%"
        cur.execute(
            """
            SELECT
                wo.id,
                wo.created_at,
                wo.order_no,
                wo.description,
                wo.status,
                wo.unit_id,
                u.name AS unit_name
            FROM work_orders wo
            LEFT JOIN units u ON u.id = wo.unit_id
            WHERE
                py_lower(wo.order_no) LIKE ?
                OR py_lower(COALESCE(wo.description, '')) LIKE ?
                OR py_lower(COALESCE(u.name, '')) LIKE ?
                OR py_lower(wo.status) LIKE ?
            ORDER BY wo.id DESC
            """,
            (pattern, pattern, pattern, pattern),
        )
        return cur.fetchall()

    def add_work_order(
        self,
        order_no: str,
        unit_id: int | None,
        description: str,
        status: str,
    ) -> int:
        cur = self.conn.cursor()
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            self.conn.execute("BEGIN")
            cur.execute(
                """
                INSERT INTO work_orders(created_at, order_no, unit_id, description, status)
                VALUES (?, ?, ?, ?, ?)
                """,
                (created_at, order_no.strip(), unit_id, description.strip(), status),
            )
            wid = cur.lastrowid
            self._audit("WORK_ORDER_INSERT", "work_orders", wid, f"order_no={order_no.strip()} status={status}", commit=False)
            self.conn.commit()
            return wid
        except Exception:
            self.conn.rollback()
            raise

    def update_work_order(
        self,
        work_order_id: int,
        order_no: str,
        unit_id: int | None,
        description: str,
        status: str,
    ):
        cur = self.conn.cursor()
        try:
            self.conn.execute("BEGIN")
            cur.execute(
                """
                UPDATE work_orders
                SET order_no = ?, unit_id = ?, description = ?, status = ?
                WHERE id = ?
                """,
                (order_no.strip(), unit_id, description.strip(), status, work_order_id),
            )
            self._audit("WORK_ORDER_UPDATE", "work_orders", work_order_id, f"order_no={order_no.strip()} status={status}", commit=False)
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def delete_work_order(self, work_order_id: int):
        try:
            self.conn.execute("BEGIN")
            cur = self.conn.cursor()
            cur.execute("DELETE FROM work_order_items WHERE work_order_id = ?", (work_order_id,))
            cur.execute("DELETE FROM work_orders WHERE id = ?", (work_order_id,))
            self._audit("WORK_ORDER_DELETE", "work_orders", work_order_id, "", commit=False)
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def get_work_order_items(self, work_order_id: int):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                woi.id,
                woi.work_order_id,
                woi.variant_id,
                woi.quantity AS requested_qty,
                v.full_code,
                v.size_name,
                i.name AS item_name,
                i.type AS item_type
            FROM work_order_items woi
            JOIN variants v ON v.id = woi.variant_id
            JOIN items i ON i.id = v.item_id
            WHERE woi.work_order_id = ?
            ORDER BY i.name COLLATE NOCASE, v.size_name COLLATE NOCASE
            """,
            (work_order_id,),
        )
        return cur.fetchall()

    def add_work_order_item(self, work_order_id: int, variant_id: int, quantity: int):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id, quantity FROM work_order_items WHERE work_order_id = ? AND variant_id = ?",
            (work_order_id, variant_id),
        )
        row = cur.fetchone()
        if row:
            cur.execute(
                "UPDATE work_order_items SET quantity = ? WHERE id = ?",
                (int(row["quantity"]) + int(quantity), row["id"]),
            )
            item_id = row["id"]
        else:
            cur.execute(
                "INSERT INTO work_order_items(work_order_id, variant_id, quantity) VALUES (?, ?, ?)",
                (work_order_id, variant_id, int(quantity)),
            )
            item_id = cur.lastrowid
        self.conn.commit()
        self._audit("WORK_ORDER_ITEM_UPSERT", "work_order_items", item_id, f"work_order_id={work_order_id} variant_id={variant_id} qty={quantity}")
        self.recompute_work_order_status(work_order_id)
        return item_id

    def update_work_order_item_qty(self, item_id: int, quantity: int):
        cur = self.conn.cursor()
        cur.execute("SELECT work_order_id FROM work_order_items WHERE id = ?", (item_id,))
        row = cur.fetchone()
        if not row:
            return
        work_order_id = row["work_order_id"]
        cur.execute("UPDATE work_order_items SET quantity = ? WHERE id = ?", (int(quantity), item_id))
        self.conn.commit()
        self._audit("WORK_ORDER_ITEM_UPDATE", "work_order_items", item_id, f"qty={quantity}")
        self.recompute_work_order_status(work_order_id)

    def delete_work_order_item(self, item_id: int):
        cur = self.conn.cursor()
        cur.execute("SELECT work_order_id FROM work_order_items WHERE id = ?", (item_id,))
        row = cur.fetchone()
        if not row:
            return
        work_order_id = row["work_order_id"]
        cur.execute("DELETE FROM work_order_items WHERE id = ?", (item_id,))
        self.conn.commit()
        self._audit("WORK_ORDER_ITEM_DELETE", "work_order_items", item_id, "")
        self.recompute_work_order_status(work_order_id)

    def get_work_order_item_issue_stats(self, work_order_id: int):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                j.variant_id,
                SUM(
                    CASE
                        WHEN i.type = 'qty' THEN COALESCE(j.quantity, 0)
                        ELSE 1
                    END
                ) AS issued_qty
            FROM journal j
            JOIN variants v ON v.id = j.variant_id
            JOIN items i ON i.id = v.item_id
            WHERE j.op_type = 'OUT' AND j.work_order_id = ?
            GROUP BY j.variant_id
            """,
            (work_order_id,),
        )
        return {r["variant_id"]: int(r["issued_qty"] or 0) for r in cur.fetchall()}

    def get_work_order_issue_documents(self, work_order_id: int) -> list[dict]:
        """Уникальные номера документов ВЫДАЧА по наряду и период дат по журналу."""
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                TRIM(j.doc_name) AS doc_name,
                MIN(j.date) AS first_date,
                MAX(j.date) AS last_date,
                COUNT(*) AS op_lines
            FROM journal j
            WHERE j.work_order_id = ?
              AND j.op_type = 'OUT'
              AND j.doc_name IS NOT NULL
              AND TRIM(j.doc_name) != ''
            GROUP BY TRIM(j.doc_name)
            ORDER BY first_date DESC, doc_name COLLATE NOCASE
            """,
            (work_order_id,),
        )
        return [dict(r) for r in cur.fetchall()]

    def recompute_work_order_status(self, work_order_id: int):
        items = self.get_work_order_items(work_order_id)
        if not items:
            new_status = "не реализован"
        else:
            issued_by_variant = self.get_work_order_item_issue_stats(work_order_id)
            requested_total = sum(int(r["requested_qty"] or 0) for r in items)
            issued_total = 0
            for row in items:
                req = int(row["requested_qty"] or 0)
                issued = int(issued_by_variant.get(row["variant_id"], 0))
                issued_total += min(req, max(issued, 0))
            if issued_total <= 0:
                new_status = "не реализован"
            elif issued_total >= requested_total:
                new_status = "реализован"
            else:
                new_status = "реализован частично"
        cur = self.conn.cursor()
        cur.execute("UPDATE work_orders SET status = ? WHERE id = ?", (new_status, work_order_id))
        self.conn.commit()
        return new_status

    # --- Search & Info ---

    def search_variants(self, text: str, only_in_stock: bool = False):
        pattern = f"%{text.strip().lower()}%"
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                v.id AS variant_id,
                v.full_code,
                v.size_name,
                i.id AS item_id,
                i.name AS item_name,
                i.type AS item_type,
                i.uom AS uom,
                CASE
                    WHEN i.type = 'qty' THEN COALESCE(sq.quantity, 0)
                    ELSE (SELECT COUNT(*) FROM stock_serial ss WHERE ss.variant_id = v.id)
                END AS stock_value
            FROM variants v
            JOIN items i ON v.item_id = i.id
            LEFT JOIN stock_qty sq ON sq.variant_id = v.id
            WHERE py_lower(v.full_code) LIKE ?
               OR py_lower(i.name) LIKE ?
               OR py_lower(COALESCE(v.size_name, '')) LIKE ?
            ORDER BY i.name COLLATE NOCASE, v.size_name COLLATE NOCASE
            """,
            (pattern, pattern, pattern),
        )
        rows = cur.fetchall()
        if only_in_stock:
            rows = [r for r in rows if (r["stock_value"] or 0) > 0]
        return rows

    def get_variant_with_item(self, variant_id: int):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                v.id AS variant_id,
                v.full_code,
                v.size_name,
                i.id AS item_id,
                i.name AS item_name,
                i.base_code,
                i.uom,
                i.type AS item_type
            FROM variants v
            JOIN items i ON v.item_id = i.id
            WHERE v.id = ?
            """,
            (variant_id,),
        )
        return cur.fetchone()

    # --- Stock operations ---

    def get_qty_stock(self, variant_id: int) -> int:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT quantity FROM stock_qty WHERE variant_id = ?", (variant_id,)
        )
        row = cur.fetchone()
        return int(row["quantity"]) if row else 0

    def adjust_qty_stock(self, variant_id: int, delta: int, *, commit: bool = True):
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO stock_qty(variant_id, quantity)
            VALUES (?, ?)
            ON CONFLICT(variant_id) DO UPDATE SET quantity = quantity + excluded.quantity
            """,
            (variant_id, delta),
        )
        if commit:
            self.conn.commit()

    def serial_exists(self, factory_sn: str) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id FROM stock_serial WHERE factory_sn = ?", (factory_sn.strip(),)
        )
        return cur.fetchone() is not None

    def serial_exists_for_variant(self, variant_id: int, factory_sn: str) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id FROM stock_serial WHERE variant_id = ? AND factory_sn = ?",
            (variant_id, factory_sn.strip()),
        )
        return cur.fetchone() is not None

    def add_serial(
        self,
        variant_id: int,
        factory_sn: str,
        manufacture_year: int | None = None,
        *,
        commit: bool = True,
    ):
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO stock_serial(variant_id, factory_sn, manufacture_year)
            VALUES (?, ?, ?)
            """,
            (variant_id, factory_sn.strip(), manufacture_year),
        )
        if commit:
            self.conn.commit()

    def remove_serial(self, variant_id: int, factory_sn: str, *, commit: bool = True) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            "DELETE FROM stock_serial WHERE variant_id = ? AND factory_sn = ?",
            (variant_id, factory_sn.strip()),
        )
        deleted = cur.rowcount > 0
        if commit:
            self.conn.commit()
        return deleted

    def get_serial_manufacture_year(self, variant_id: int, factory_sn: str) -> int | None:
        """Год выпуска на складе для S/N (до списания при выдаче)."""
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT manufacture_year FROM stock_serial
            WHERE variant_id = ? AND factory_sn = ?
            """,
            (variant_id, factory_sn.strip()),
        )
        row = cur.fetchone()
        if not row or row["manufacture_year"] is None:
            return None
        return int(row["manufacture_year"])

    def update_serial_manufacture_year(
        self,
        variant_id: int,
        factory_sn: str,
        manufacture_year: int | None,
        *,
        commit: bool = True,
    ) -> bool:
        """Обновить год выпуска у единицы на складе. None — сбросить год."""
        cur = self.conn.cursor()
        cur.execute(
            """
            UPDATE stock_serial SET manufacture_year = ?
            WHERE variant_id = ? AND factory_sn = ?
            """,
            (manufacture_year, variant_id, factory_sn.strip()),
        )
        updated = cur.rowcount > 0
        if commit:
            self.conn.commit()
        return updated

    # --- Journal ---

    def _audit(self, action: str, entity_type: str, entity_id: int | None, details: str = "", *, commit: bool = True):
        """Write to audit_log. Set commit=False when called inside an outer transaction."""
        try:
            cur = self.conn.cursor()
            cur.execute(
                """
                INSERT INTO audit_log(ts, action, entity_type, entity_id, details)
                VALUES (?, ?, ?, ?, ?)
                """,
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), action, entity_type, entity_id or 0, details),
            )
            if commit:
                self.conn.commit()
        except Exception as e:
            logger.exception("Audit log write failed: %s", e)

    def add_journal_record(
        self,
        op_type: str,
        variant_id: int,
        quantity: int | None,
        factory_sn: str | None,
        unit_id: int | None,
        doc_name: str,
        *,
        work_order_id: int | None = None,
        manufacture_year: int | None = None,
        commit: bool = True,
    ):
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO journal(
                date, op_type, variant_id, quantity, factory_sn, manufacture_year,
                unit_id, doc_name, work_order_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                op_type,
                variant_id,
                quantity,
                factory_sn.strip() if factory_sn else None,
                manufacture_year,
                unit_id,
                doc_name.strip(),
                work_order_id,
            ),
        )
        jid = cur.lastrowid
        self._audit("JOURNAL_INSERT", "journal", jid, f"op={op_type} variant_id={variant_id} doc={doc_name.strip()}", commit=False)
        if commit:
            self.conn.commit()

    def post_operation(self, basket: list[dict], op_type: str, unit_id: int, doc_name: str,
                       work_order_id: int | None = None):
        """Atomically post all basket items in a single transaction."""
        try:
            self.conn.execute("BEGIN")
            for pos in basket:
                vid = pos["variant_id"]
                if pos["item_type"] == "qty":
                    delta = pos["qty"] if op_type == "IN" else -pos["qty"]
                    self.adjust_qty_stock(vid, delta, commit=False)
                    self.add_journal_record(op_type, vid, pos["qty"], None, unit_id, doc_name,
                                            work_order_id=work_order_id, commit=False)
                else:
                    sn = pos["sn"]
                    if op_type == "IN":
                        my = pos.get("manufacture_year")
                        if my is not None:
                            my = int(my)
                        self.add_serial(vid, sn, manufacture_year=my, commit=False)
                        self.add_journal_record(
                            op_type, vid, 1, sn, unit_id, doc_name,
                            work_order_id=work_order_id, manufacture_year=my, commit=False,
                        )
                    else:
                        my = self.get_serial_manufacture_year(vid, sn)
                        self.remove_serial(vid, sn, commit=False)
                        self.add_journal_record(
                            op_type, vid, 1, sn, unit_id, doc_name,
                            work_order_id=work_order_id, manufacture_year=my, commit=False,
                        )
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def reverse_operation(self, journal_rows: list[dict]):
        """Atomically reverse a group of journal entries (one document).

        For each row:
          - OUT + qty  → add quantity back to stock
          - OUT + serial → re-add serial to stock
          - IN  + qty  → subtract quantity from stock
          - IN  + serial → remove serial from stock
        Then delete the journal rows.
        """
        if not journal_rows:
            return
        try:
            self.conn.execute("BEGIN")
            wo_ids = set()
            for r in journal_rows:
                jid = r["id"]
                vid = r["variant_id"]
                op = r["op_type"]
                item_type = r["item_type"]
                qty = int(r["quantity"] or 0)
                sn = r["factory_sn"]
                try:
                    wo_id = r["work_order_id"]
                except (KeyError, IndexError):
                    wo_id = None
                if wo_id:
                    wo_ids.add(wo_id)

                if item_type == "qty":
                    delta = -qty if op == "IN" else qty
                    self.adjust_qty_stock(vid, delta, commit=False)
                else:
                    if op == "IN":
                        self.remove_serial(vid, sn, commit=False)
                    else:
                        if self.serial_exists_for_variant(vid, sn):
                            raise ValueError(f"S/N «{sn}» уже на складе — возможно, операция уже была отменена")
                        my_rev = r["manufacture_year"]
                        my_i = int(my_rev) if my_rev is not None else None
                        self.add_serial(vid, sn, manufacture_year=my_i, commit=False)

                self.conn.execute("DELETE FROM journal WHERE id = ?", (jid,))
                self._audit("JOURNAL_REVERSE", "journal", jid,
                            f"reversed op={op} variant_id={vid}", commit=False)

            self.conn.commit()
            for wo_id in wo_ids:
                self.recompute_work_order_status(wo_id)
        except Exception:
            self.conn.rollback()
            raise

    # --- Stock view ---

    def get_stock_view(self):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                i.id AS item_id,
                i.name AS item_name,
                i.base_code,
                i.type AS item_type,
                i.uom AS uom,
                COALESCE(
                    CASE
                        WHEN i.type = 'qty' THEN SUM(sq.quantity)
                        ELSE COUNT(ss.id)
                    END,
                    0
                ) AS stock_value
            FROM items i
            LEFT JOIN variants v ON v.item_id = i.id
            LEFT JOIN stock_qty sq ON sq.variant_id = v.id
            LEFT JOIN stock_serial ss ON ss.variant_id = v.id
            GROUP BY i.id, i.name, i.base_code, i.type, i.uom
            HAVING stock_value > 0
            ORDER BY i.name COLLATE NOCASE
            """
        )
        return cur.fetchall()

    @staticmethod
    def _journal_sql_filters(
        date_from: str | None,
        date_to: str | None,
        unit_id: int | None,
    ) -> tuple[str, list]:
        where_parts: list[str] = []
        params: list = []
        if date_from:
            where_parts.append("date(j.date) >= ?")
            params.append(date_from)
        if date_to:
            where_parts.append("date(j.date) <= ?")
            params.append(date_to)
        if unit_id is not None:
            where_parts.append("j.unit_id = ?")
            params.append(unit_id)
        where_sql = " WHERE " + " AND ".join(where_parts) if where_parts else ""
        return where_sql, params

    def count_journal_rows(
        self,
        date_from: str | None = None,
        date_to: str | None = None,
        unit_id: int | None = None,
    ) -> int:
        """Число строк журнала с теми же фильтрами, что у get_journal_view (без лимита)."""
        where_sql, params = self._journal_sql_filters(date_from, date_to, unit_id)
        cur = self.conn.cursor()
        cur.execute(f"SELECT COUNT(*) AS cnt FROM journal j {where_sql}", tuple(params))
        row = cur.fetchone()
        return int(row["cnt"] if row else 0)

    def get_journal_view(
        self,
        limit: int = 200,
        date_from: str | None = None,
        date_to: str | None = None,
        unit_id: int | None = None,
    ):
        cur = self.conn.cursor()
        where_sql, params = self._journal_sql_filters(date_from, date_to, unit_id)
        params_with_limit = list(params) + [limit]
        cur.execute(
            f"""
            SELECT
                j.id AS id,
                j.date AS date,
                j.op_type AS op_type,
                j.doc_name AS doc_name,
                j.variant_id AS variant_id,
                j.work_order_id AS work_order_id,
                wo.order_no AS work_order_no,
                v.size_name AS size_name,
                v.full_code AS full_code,
                i.name AS item_name,
                i.type AS item_type,
                j.quantity AS quantity,
                j.factory_sn AS factory_sn,
                j.manufacture_year AS manufacture_year,
                u.name AS unit_name
            FROM journal j
            JOIN variants v ON j.variant_id = v.id
            JOIN items i ON v.item_id = i.id
            LEFT JOIN units u ON j.unit_id = u.id
            LEFT JOIN work_orders wo ON j.work_order_id = wo.id
            {where_sql}
            ORDER BY j.date DESC, j.id DESC
            LIMIT ?
            """,
            tuple(params_with_limit),
        )
        return cur.fetchall()

    def get_serials_for_variant(self, variant_id: int):
        """Возвращает все серийные номера на складе для данного варианта."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT factory_sn FROM stock_serial WHERE variant_id = ? ORDER BY factory_sn COLLATE NOCASE",
            (variant_id,),
        )
        return cur.fetchall()

    def get_serials_for_item(self, item_id: int):
        """Возвращает все S/N на складе по всем вариантам изделия (с variant_id для корзины)."""
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT ss.factory_sn, ss.manufacture_year AS manufacture_year,
                   v.id AS variant_id, v.size_name, v.full_code
            FROM stock_serial ss
            JOIN variants v ON v.id = ss.variant_id
            WHERE v.item_id = ?
            ORDER BY v.size_name COLLATE NOCASE, ss.factory_sn COLLATE NOCASE
            """,
            (item_id,),
        )
        return cur.fetchall()

    def get_item_stock_details(self, item_id: int):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                v.id AS variant_id,
                v.full_code,
                v.size_name,
                i.uom AS uom,
                COALESCE(
                    CASE
                        WHEN i.type = 'qty' THEN sq.quantity
                        ELSE COUNT(ss.id)
                    END,
                    0
                ) AS stock_value
            FROM variants v
            JOIN items i ON v.item_id = i.id
            LEFT JOIN stock_qty sq ON sq.variant_id = v.id
            LEFT JOIN stock_serial ss ON ss.variant_id = v.id
            WHERE v.item_id = ?
            GROUP BY v.id, v.full_code, v.size_name, i.uom, sq.quantity
            ORDER BY v.size_name COLLATE NOCASE
            """,
            (item_id,),
        )
        return cur.fetchall()


# Максимум строк журнала в одном файле Excel/PDF (самые новые по дате).
JOURNAL_EXPORT_ROW_LIMIT = 10000


# --- Экспорт отчётов: Excel / PDF ---

def _export_journal_excel(db: DatabaseManager, path: str, date_from: str, date_to: str, unit_id: int | None) -> bool:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        logger.warning("openpyxl not installed: pip install openpyxl")
        return False
    try:
        rows = db.get_journal_view(
            limit=JOURNAL_EXPORT_ROW_LIMIT,
            date_from=date_from,
            date_to=date_to,
            unit_id=unit_id,
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Журнал операций"
        headers = [
            "Дата",
            "Операция",
            "Документ",
            "Наряд",
            "Название",
            "Размер",
            "Кол-во",
            "Год выпуска",
            "Подразделение",
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
            ws.cell(1, c).font = Font(bold=True)
        for r, row in enumerate(rows, 2):
            op_text = "Приход" if row["op_type"] == "IN" else "Выдача"
            qty = row["quantity"] if row["quantity"] is not None else (1 if row["factory_sn"] else 0)
            my = row["manufacture_year"]
            my_cell = int(my) if my is not None else ""
            ws.cell(r, 1, row["date"])
            ws.cell(r, 2, op_text)
            ws.cell(r, 3, row["doc_name"] or "")
            ws.cell(r, 4, (row["work_order_no"] or "").strip())
            ws.cell(r, 5, row["item_name"])
            ws.cell(r, 6, row["size_name"])
            ws.cell(r, 7, qty)
            ws.cell(r, 8, my_cell)
            ws.cell(r, 9, row["unit_name"] or "")
        wb.save(path)
        logger.info("Exported journal to Excel: %s", path)
        return True
    except Exception as e:
        logger.exception("Export journal Excel failed: %s", e)
        return False


def _register_pdf_font() -> str:
    """
    Регистрирует Arial из системных шрифтов Windows для поддержки кириллицы.
    Возвращает имя зарегистрированного шрифта.
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "ArialUnicode"
    try:
        pdfmetrics.getFont(font_name)
        return font_name
    except Exception as e:
        logger.debug("PDF getFont(%s): %s", font_name, e)

    candidates = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/Arial.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont(font_name, path))
            return font_name

    return "Helvetica"


def _export_journal_pdf(db: DatabaseManager, path: str, date_from: str, date_to: str, unit_id: int | None) -> bool:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
    except ImportError:
        logger.warning("reportlab not installed: pip install reportlab")
        return False
    try:
        font = _register_pdf_font()
        rows = db.get_journal_view(
            limit=JOURNAL_EXPORT_ROW_LIMIT,
            date_from=date_from,
            date_to=date_to,
            unit_id=unit_id,
        )

        doc = SimpleDocTemplate(
            path,
            pagesize=landscape(A4),
            leftMargin=10 * mm, rightMargin=10 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
        )

        title_style = ParagraphStyle("title", fontName=font, fontSize=12, spaceAfter=6)
        header = Paragraph(
            f"Журнал операций ({date_from} — {date_to})",
            title_style,
        )

        col_headers = [
            "Дата",
            "Операция",
            "Документ",
            "Наряд",
            "Название",
            "Размер",
            "Кол-во",
            "Год вып.",
            "Подразделение",
        ]
        data = [col_headers]
        for row in rows:
            op_text = "Приход" if row["op_type"] == "IN" else "Выдача"
            qty = row["quantity"] if row["quantity"] is not None else (1 if row["factory_sn"] else 0)
            my = row["manufacture_year"]
            my_txt = str(int(my)) if my is not None else ""
            data.append([
                row["date"],
                op_text,
                row["doc_name"] or "",
                (row["work_order_no"] or "").strip(),
                row["item_name"],
                row["size_name"],
                str(qty),
                my_txt,
                row["unit_name"] or "",
            ])

        col_widths = [30*mm, 18*mm, 28*mm, 24*mm, 56*mm, 18*mm, 14*mm, 14*mm, 32*mm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f7aec")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), font),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4ff")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d5e8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ]))

        doc.build([header, Spacer(1, 4*mm), t])
        logger.info("Exported journal to PDF: %s", path)
        return True
    except Exception as e:
        logger.exception("Export journal PDF failed: %s", e)
        return False


def _export_stock_excel(db: DatabaseManager, path: str) -> bool:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        logger.warning("openpyxl not installed: pip install openpyxl")
        return False
    try:
        rows = db.get_stock_view()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Остатки"
        headers = ["Н/Н (базовый)", "Название", "Остаток", "Ед. изм."]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
            ws.cell(1, c).font = Font(bold=True)
        for r, row in enumerate(rows, 2):
            ws.cell(r, 1, row["base_code"])
            ws.cell(r, 2, row["item_name"])
            ws.cell(r, 3, row["stock_value"])
            ws.cell(r, 4, row["uom"])
        wb.save(path)
        logger.info("Exported stock to Excel: %s", path)
        return True
    except Exception as e:
        logger.exception("Export stock Excel failed: %s", e)
        return False


def _export_stock_pdf(db: DatabaseManager, path: str) -> bool:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
    except ImportError:
        logger.warning("reportlab not installed: pip install reportlab")
        return False
    try:
        font = _register_pdf_font()

        doc = SimpleDocTemplate(
            path,
            pagesize=A4,
            leftMargin=10 * mm, rightMargin=10 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
        )

        title_style = ParagraphStyle("title", fontName=font, fontSize=12, spaceAfter=6)
        sub_style = ParagraphStyle("sub", fontName=font, fontSize=8, textColor=colors.HexColor("#555555"), spaceAfter=2)

        header = Paragraph("Остатки на складе", title_style)
        header_date = Paragraph(
            f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            sub_style,
        )

        base_rows = db.get_stock_view()

        elements = [header, header_date, Spacer(1, 4 * mm)]

        COL_W = [35*mm, 70*mm, 30*mm, 20*mm]

        for row in base_rows:
            details = db.get_item_stock_details(row["item_id"])

            item_data = [[
                row["base_code"],
                row["item_name"],
                f"Итого: {row['stock_value']}",
                row["uom"],
            ]]
            item_table = Table(item_data, colWidths=COL_W)
            item_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1f7aec")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#1764c0")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(item_table)

            if details:
                size_data = [["  Размер", "Н/Н (полный)", "Остаток", "Ед. изм."]]
                for d in details:
                    size_data.append([
                        f"  {d['size_name']}",
                        d["full_code"],
                        str(d["stock_value"]),
                        d["uom"],
                    ])
                size_table = Table(size_data, colWidths=COL_W)
                size_table.setStyle(TableStyle([
                    ("FONTNAME", (0, 0), (-1, -1), font),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8edf7")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7ff")]),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d0d5e8")),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ]))
                elements.append(size_table)

            elements.append(Spacer(1, 3 * mm))

        doc.build(elements)
        logger.info("Exported stock to PDF: %s", path)
        return True
    except Exception as e:
        logger.exception("Export stock PDF failed: %s", e)
        return False
